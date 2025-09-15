#!/usr/bin/env python3
"""Publish the NetBox VMs CSV as a Confluence table."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import html
import json
import os
import subprocess
import sys
from collections.abc import Iterable, Sequence
from pathlib import Path

import requests
from rich import print

from enreach_tools.env import get_env, load_env, project_root

CONFLUENCE_STORAGE = "storage"
SELECTED_COLUMNS = ["Name", "Status", "Cluster", "IP Address", "Device"]


def env_flag(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def read_csv(path: Path, limit: int | None = None) -> tuple[Sequence[str], list[list[str]]]:
    if not path.exists():
        print(f"[red]CSV not found:[/red] {path}")
        raise SystemExit(1)
    with path.open("r", newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration:
            print("[yellow]CSV is empty; nothing to publish[/yellow]")
            return [], []
        rows: list[list[str]] = []
        for idx, row in enumerate(reader):
            if limit is not None and idx >= limit:
                break
            rows.append(row)
    return header, rows


def load_column_order(limit_column: str | None = None) -> list[str]:
    order_path = project_root() / "netbox-export/etc/column_order.xlsx"
    if not order_path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        print(f"[yellow]Warning:[/yellow] unable to read column order ({exc}); using CSV header order")
        return []
    wb = load_workbook(order_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            cols: list[str] = []
            for cell in row:
                if cell is None:
                    continue
                name = str(cell).strip()
                if not name:
                    continue
                cols.append(name)
                if limit_column and name == limit_column:
                    return cols
            return cols
    finally:
        wb.close()
    return []


def apply_column_subset(header: Sequence[str], rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    ordered = load_column_order()
    selected: list[str] = []
    for name in ordered:
        if name in SELECTED_COLUMNS and name in header and name not in selected:
            selected.append(name)
    for name in SELECTED_COLUMNS:
        if name in header and name not in selected:
            selected.append(name)
    if not selected:
        return list(header), rows

    indexes = [header.index(name) for name in selected]
    trimmed_rows: list[list[str]] = []
    for row in rows:
        trimmed = []
        for idx in indexes:
            trimmed.append(row[idx] if idx < len(row) else "")
        trimmed_rows.append(trimmed)
    return selected, trimmed_rows


def escape_cell(value: str) -> str:
    if value is None:
        return ""
    return html.escape(value, quote=False)


def build_table_html(header: Sequence[str], rows: Iterable[Sequence[str]]) -> str:
    if not header:
        return "<p><em>No data</em></p>"
    head_cells = "".join(f"<th>{escape_cell(col)}</th>" for col in header)
    body_rows = []
    for row in rows:
        cells = [escape_cell(col) for col in row]
        if len(cells) < len(header):
            cells += [""] * (len(header) - len(cells))
        cells_html = "".join(f"<td>{cell}</td>" for cell in cells[: len(header)])
        body_rows.append(f"<tr>{cells_html}</tr>")
    body_html = "".join(body_rows)
    table_html = (
        "<table data-layout=\"wide\" style=\"width:100%\">"
        "<thead><tr>" + head_cells + "</tr></thead>"
        "<tbody>" + body_html + "</tbody>"
        "</table>"
    )
    return table_html


def wrap_with_macros(table_html: str, enable_filter: bool, enable_sort: bool) -> str:
    content = table_html
    if enable_sort:
        content = (
            '<ac:structured-macro ac:name="table-sort" ac:schema-version="1">'
            "<ac:rich-text-body>"
            f"{content}"
            "</ac:rich-text-body>"
            "</ac:structured-macro>"
        )
    if enable_filter:
        content = (
            '<ac:structured-macro ac:name="table-filter" ac:schema-version="1">'
            "<ac:rich-text-body>"
            f"{content}"
            "</ac:rich-text-body>"
            "</ac:structured-macro>"
        )
    return content


def build_page_body(attachments_macro: str, heading: str, table_html: str, note: str | None = None) -> str:
    parts = [attachments_macro, "<p />"]
    if heading:
        parts.append(f"<h2>{escape_cell(heading)}</h2>")
    if note:
        parts.append(f"<p><em>{escape_cell(note)}</em></p>")
    parts.append(table_html)
    return "".join(parts)


def fetch_page(base: str, auth: tuple[str, str], page_id: str) -> dict:
    url = f"{base}/rest/api/content/{page_id}?expand=body.storage,version,title"
    resp = requests.get(url, auth=auth, timeout=30)
    if resp.status_code != 200:
        print(f"[red]Failed to fetch page {page_id} ({resp.status_code}):[/red] {resp.text[:400]}")
        raise SystemExit(1)
    return resp.json()


def main() -> int:
    load_env()

    parser = argparse.ArgumentParser(description="Publish NetBox VMs CSV as Confluence table")
    parser.add_argument(
        "--csv",
        default="netbox-export/data/netbox_vms_export.csv",
        help="Path to the VMs export CSV",
    )
    parser.add_argument(
        "--page-id",
        default=os.getenv("CONFLUENCE_VMS_PAGE_ID")
        or os.getenv("CONFLUENCE_PAGE_ID"),
        help="Target Confluence page ID",
    )
    parser.add_argument(
        "--heading",
        default="NetBox VMs Export",
        help="Heading placed above the table",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit number of data rows (omit for full table)",
    )
    default_filter = env_flag("CONFLUENCE_ENABLE_TABLE_FILTER", False)
    default_sort = env_flag("CONFLUENCE_ENABLE_TABLE_SORT", False)
    parser.add_argument(
        "--filter",
        dest="filter_macro",
        action="store_true",
        default=default_filter,
        help="Wrap table in the Table Filter macro (requires Table Filter & Charts app)",
    )
    parser.add_argument(
        "--no-filter",
        dest="filter_macro",
        action="store_false",
        help="Do not wrap table in the Table Filter macro",
    )
    parser.add_argument(
        "--sort",
        dest="sort_macro",
        action="store_true",
        default=default_sort,
        help="Wrap table in the Table Sort macro",
    )
    parser.add_argument(
        "--no-sort",
        dest="sort_macro",
        action="store_false",
        help="Do not wrap table in the Table Sort macro",
    )
    parser.add_argument(
        "--message",
        default="Updated NetBox VMs table",
        help="Confluence version comment",
    )
    parser.add_argument(
        "--minor",
        dest="minor",
        action="store_true",
        help="Mark the update as a minor edit",
    )
    parser.add_argument(
        "--major",
        dest="minor",
        action="store_false",
        help="Mark the update as a major edit",
    )
    parser.set_defaults(filter_macro=default_filter, sort_macro=default_sort, minor=True)
    args = parser.parse_args()

    if not args.page_id:
        print("[red]Missing page ID[/red]. Supply --page-id or set CONFLUENCE_VMS_PAGE_ID.")
        return 2

    root = project_root()
    csv_path = root / args.csv if not os.path.isabs(args.csv) else Path(args.csv)
    header, rows = read_csv(csv_path, args.limit)
    header, rows = apply_column_subset(header, rows)
    table_html = build_table_html(header, rows)
    table_with_macros = wrap_with_macros(table_html, enable_filter=args.filter_macro, enable_sort=args.sort_macro)

    base = get_env("ATLASSIAN_BASE_URL", required=True).rstrip("/") + "/wiki"
    email = get_env("ATLASSIAN_EMAIL", required=True)
    token = get_env("ATLASSIAN_API_TOKEN", required=True)
    auth = (email, token)

    page = fetch_page(base, auth, args.page_id)

    timestamp = dt.datetime.now(dt.UTC).strftime("%Y-%m-%d %H:%M UTC")
    note = (
        "Last synced from NetBox export on {ts}. Showing columns: {cols}."
    ).format(ts=timestamp, cols=", ".join(SELECTED_COLUMNS))
    attachments_macro = '<ac:structured-macro ac:name="attachments" ac:schema-version="1" data-layout="wide" />'
    body = build_page_body(attachments_macro, args.heading, table_with_macros, note=note)

    page_id = page.get("id")
    title = page.get("title")
    version = page.get("version", {}).get("number", 1)
    url = f"{base}/rest/api/content/{page_id}"
    payload = {
        "id": page_id,
        "type": page.get("type", "page"),
        "title": title,
        "version": {"number": version + 1, "minorEdit": args.minor},
        "body": {CONFLUENCE_STORAGE: {"value": body, "representation": CONFLUENCE_STORAGE}},
    }
    if args.message:
        payload["version"]["message"] = args.message
    headers = {"Content-Type": "application/json"}
    resp = requests.put(url, auth=auth, headers=headers, data=json.dumps(payload), timeout=30)
    if resp.status_code not in (200, 202):
        print(f"[red]Failed to update page ({resp.status_code}):[/red] {resp.text[:400]}")
        raise SystemExit(1)

    upload_script = project_root() / "netbox-export/bin/confluence_upload_attachment.py"
    if upload_script.exists():
        cmd = [
            sys.executable,
            str(upload_script),
            "--file",
            str(csv_path),
            "--page-id",
            args.page_id,
            "--name",
            csv_path.name,
        ]
        if args.message:
            cmd += ["--comment", args.message]
        code = subprocess.call(cmd, cwd=project_root(), env=os.environ.copy())
        if code != 0:
            print(f"[yellow]Warning:[/yellow] uploading CSV attachment failed (exit {code})")

    print("[green]Confluence page updated with NetBox VMs table[/green]")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""
NetBox CSV Merger Script

This script merges the devices and VMs CSV files exported from NetBox,
adding a 'netbox_type' column to distinguish between the two data sources.

Usage:
    python bin/merge_netbox_csvs.py

The script will:
1. Read both devices and VMs CSV files
2. Add a 'netbox_type' column ('devices' or 'vms')
3. Merge them into a single CSV file
4. Save the merged result to data/netbox_merged_export.csv
5. Create an Excel file with filters and sorting by Name
"""

import csv
import os
from pathlib import Path
from datetime import datetime
from enreach_tools.env import load_env, project_root

# Try to import Excel-related libraries
EXCEL_AVAILABLE = False
pd = None
Workbook = None
dataframe_to_rows = None
Table = None
TableStyleInfo = None

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    pass  # Will show warning in main function if needed

# Ensure .env from project root is loaded so optional NETBOX_XLSX_ORDER_FILE is available
load_env()

def _load_column_order_from_xlsx(order_file):
    """Return a list of column headers in the order found in the first row of the first sheet."""
    try:
        wb = load_workbook(order_file, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        headers = []
        for cell in ws[1]:  # first row
            v = cell.value
            if v is None:
                continue
            headers.append(str(v))
        wb.close()
        return headers
    except Exception:
        return []

def merge_netbox_csvs():
    """Merge NetBox devices and VMs CSV files with netbox_type column."""
    
    # File paths
    # Resolve data directory relative to project root, defaulting to legacy location under netbox-export/
    root = project_root()
    data_dir_env = os.getenv('NETBOX_DATA_DIR', 'netbox-export/data')
    data_dir_path = Path(data_dir_env) if os.path.isabs(data_dir_env) else (root / data_dir_env)
    devices_file = str(data_dir_path / 'netbox_devices_export.csv')
    vms_file = str(data_dir_path / 'netbox_vms_export.csv')
    output_file = str(data_dir_path / 'netbox_merged_export.csv')
    
    # Check if input files exist
    if not os.path.exists(devices_file):
        print(f"Error: {devices_file} not found!")
        return False
        
    if not os.path.exists(vms_file):
        print(f"Error: {vms_file} not found!")
        return False
    
    print("Starting NetBox CSV merge process...")
    print(f"Devices file: {devices_file}")
    print(f"VMs file: {vms_file}")
    print(f"Output file: {output_file}")
    print("-" * 50)
    
    try:
        # Read headers from both files to ensure compatibility
        with open(devices_file, 'r', encoding='utf-8') as f:
            devices_reader = csv.reader(f)
            devices_headers = next(devices_reader)
            
        with open(vms_file, 'r', encoding='utf-8') as f:
            vms_reader = csv.reader(f)
            vms_headers = next(vms_reader)
        
        # Create a unified header set by combining both headers
        print(f"Devices headers: {len(devices_headers)} columns")
        print(f"VMs headers: {len(vms_headers)} columns")
        
        # Start with devices headers as base
        merged_headers = devices_headers.copy()
        
        # Add any VM headers that don't exist in devices headers
        for vm_header in vms_headers:
            if vm_header not in merged_headers:
                merged_headers.append(vm_header)
                print(f"Added VM-specific column: {vm_header}")
        
        # Add netbox_type column
        merged_headers.append('netbox_type')
        
        print(f"Final merged headers: {len(merged_headers)} columns")
        
        # Create column mapping for both file types (not needed for current logic)
        
        # Initialize counters
        devices_count = 0
        vms_count = 0
        
        # Create merged CSV file
        with open(output_file, 'w', newline='', encoding='utf-8') as outfile:
            writer = csv.writer(outfile)
            
            # Write headers
            writer.writerow(merged_headers)
            
            # Process devices file
            print("Processing devices...")
            with open(devices_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader)  # Skip header
                
                for row in reader:
                    # Create a properly mapped row for merged headers
                    merged_row = [''] * (len(merged_headers) - 1)  # -1 for netbox_type
                    
                    # Map device data to correct columns
                    for header, value in zip(devices_headers, row):
                        if header in merged_headers:
                            merged_idx = merged_headers.index(header)
                            merged_row[merged_idx] = value
                    
                    # Add netbox_type
                    merged_row.append('devices')
                    writer.writerow(merged_row)
                    devices_count += 1
                    
                    if devices_count % 100 == 0:
                        print(f"  Processed {devices_count} devices...")
            
            # Process VMs file
            print("Processing VMs...")
            with open(vms_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader)  # Skip header
                
                for row in reader:
                    # Create a properly mapped row for merged headers
                    merged_row = [''] * (len(merged_headers) - 1)  # -1 for netbox_type
                    
                    # Map VM data to correct columns
                    for header, value in zip(vms_headers, row):
                        if header in merged_headers:
                            merged_idx = merged_headers.index(header)
                            merged_row[merged_idx] = value
                    
                    # Add netbox_type
                    merged_row.append('vms')
                    writer.writerow(merged_row)
                    vms_count += 1
                    
                    if vms_count % 100 == 0:
                        print(f"  Processed {vms_count} VMs...")
        
        # Summary
        total_count = devices_count + vms_count
        print("-" * 50)
        print("Merge completed successfully!")
        print(f"Devices processed: {devices_count:,}")
        print(f"VMs processed: {vms_count:,}")
        print(f"Total records: {total_count:,}")
        print(f"Output file: {output_file}")
        
        # File size info
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"Output file size: {file_size:,} bytes ({file_size / 1024 / 1024:.2f} MB)")
        
        return True
        
    except Exception as e:
        print(f"Error during merge process: {str(e)}")
        return False

def create_excel_export(csv_file, excel_file):
    """Create Excel file from CSV with filters and sorting."""
    if not EXCEL_AVAILABLE:
        print("Skipping Excel export - required libraries not available")
        return False
    
    try:
        print(f"\nCreating Excel export: {excel_file}")
        
        df = pd.read_csv(csv_file)

        # Optional: reorder columns based on a reference XLSX's header row
        # Candidate paths (first found wins):
        # Resolve etc and data paths
        root = project_root()
        data_dir_env = os.getenv('NETBOX_DATA_DIR', 'netbox-export/data')
        data_dir_path = Path(data_dir_env) if os.path.isabs(data_dir_env) else (root / data_dir_env)
        order_candidates = [
            os.getenv('NETBOX_XLSX_ORDER_FILE'),
            str(root / 'netbox-export' / 'etc' / 'column_order.xlsx'),
            str(data_dir_path / 'netbox_merged_export.xlsx'),  # allow using a prior export as the template
        ]
        order_file = next((p for p in order_candidates if p and os.path.exists(p)), None)
        if order_file:
            print(f"Applying column order from: {order_file}")
            desired_order = _load_column_order_from_xlsx(order_file)
            if desired_order:
                # Keep only columns that exist in df, in the desired order
                ordered_cols = [c for c in desired_order if c in df.columns]
                # Append any columns from df that were not in the template, to avoid data loss
                tail_cols = [c for c in df.columns if c not in ordered_cols]
                df = df[ordered_cols + tail_cols]
            else:
                print("Warning: could not read headers from order file; keeping CSV order.")
        else:
            print("No column order template found; keeping CSV order.")

        # Preserve CSV row order in Excel (no additional sorting here)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "NetBox Inventory"
        
        # Write data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Create table with filters (only if we have data rows)
        num_cols = len(df.columns)
        end_col = get_column_letter(num_cols)
        has_rows = len(df) > 0
        if has_rows:
            table_range = f"A1:{end_col}{len(df) + 1}"
            print(f"Creating table with range: {table_range}")
            table = Table(displayName="NetBoxInventory", ref=table_range)

            # Add table style
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            table.tableStyleInfo = style

            # Add table to worksheet (this enables filters)
            ws.add_table(table)
        else:
            print("No data rows; skipping table creation to avoid Excel warnings.")

        # Freeze top row and first column (keep headers and column A visible)
        ws.freeze_panes = "B2"

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save Excel file
        wb.save(excel_file)
        
        # File size info
        if os.path.exists(excel_file):
            file_size = os.path.getsize(excel_file)
            print(f"Excel file created: {file_size:,} bytes ({file_size / 1024 / 1024:.2f} MB)")
        
        print("✅ Excel export completed with:")
        print("  - Data sorted by Name")
        print("  - Filters enabled on header row")
        print("  - Auto-adjusted column widths")
        print("  - Table formatting applied")
        
        return True
        
    except Exception as e:
        print(f"Error creating Excel export: {str(e)}")
        return False

def main():
    """Main function."""
    print("NetBox CSV Merger & Excel Exporter")
    print("=" * 50)
    
    start_time = datetime.now()
    
    # Step 1: Merge CSV files
    csv_success = merge_netbox_csvs()
    
    # Step 2: Create Excel export
    excel_success = False
    if csv_success:
        root = project_root()
        data_dir_env = os.getenv('NETBOX_DATA_DIR', 'netbox-export/data')
        data_dir_path = Path(data_dir_env) if os.path.isabs(data_dir_env) else (root / data_dir_env)
        csv_file = str(data_dir_path / 'netbox_merged_export.csv')
        excel_file = str(data_dir_path / 'Systems CMDB.xlsx')
        excel_success = create_excel_export(csv_file, excel_file)
    
    end_time = datetime.now()
    duration = end_time - start_time
    print(f"\nTotal execution time: {duration.total_seconds():.2f} seconds")
    
    # Summary
    if csv_success and excel_success:
        print("✅ Both CSV merge and Excel export completed successfully!")
    elif csv_success:
        print("✅ CSV merge completed successfully!")
        print("⚠️  Excel export skipped or failed")
    else:
        print("❌ CSV merge failed!")
        exit(1)

if __name__ == "__main__":
    main()

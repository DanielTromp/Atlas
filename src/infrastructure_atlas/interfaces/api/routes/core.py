"""Core API routes (health, configuration, logs, etc.)."""

from __future__ import annotations

import os
from pathlib import Path

from fastapi import APIRouter, Query
from fastapi.responses import FileResponse, RedirectResponse

from infrastructure_atlas.env import project_root

router = APIRouter(tags=["core"])


def _data_dir() -> Path:
    """Get the data directory path from environment."""
    root = project_root()
    raw = os.getenv("NETBOX_DATA_DIR", "data")
    path = Path(raw) if os.path.isabs(raw) else (root / raw)
    path.mkdir(parents=True, exist_ok=True)
    return path


def _csv_path(name: str) -> Path:
    """Get path to a file in the data directory."""
    return _data_dir() / name


def _chat_default_temperature() -> float | None:
    """Get default chat temperature from environment."""
    raw = os.getenv("CHAT_DEFAULT_TEMPERATURE", "").strip().lower()
    if raw in {"", "default", "auto"}:
        return None
    try:
        return float(raw)
    except Exception:
        return None


# Simple export log file (appends)
LOG_PATH = project_root() / "export.log"


@router.get("/")
def root_redirect():
    """Redirect root to the web UI."""
    return RedirectResponse(url="/app/")


@router.get("/favicon.png")
def favicon_png():
    """Serve favicon from project package location."""
    # Path: interfaces/api/routes/core.py -> need to go up to infrastructure_atlas then into api/static
    # __file__.parent = routes/, .parent.parent = interfaces/, .parent.parent.parent = infrastructure_atlas/
    base = Path(__file__).parent.parent.parent.parent  # -> infrastructure_atlas/
    static_png = base / "api" / "static" / "favicon.png"
    if static_png.exists():
        return FileResponse(static_png, media_type="image/png")
    from fastapi import HTTPException

    raise HTTPException(status_code=404, detail="favicon not found")


@router.get("/health")
def health():
    """Health check endpoint with data directory status."""
    d = _data_dir()
    dev = _csv_path("netbox_devices_export.csv")
    vms = _csv_path("netbox_vms_export.csv")
    merged = _csv_path("netbox_merged_export.csv")
    return {
        "status": "ok",
        "data_dir": str(d),
        "files": {
            "devices_csv": dev.exists(),
            "vms_csv": vms.exists(),
            "merged_csv": merged.exists(),
        },
    }


@router.get("/column-order")
def column_order() -> list[str]:
    """Return preferred column order based on Systems CMDB.xlsx if available.

    Falls back to merged CSV headers if Excel not found; otherwise empty list.
    """
    try:
        # Prefer the Excel produced by merge step
        xlsx_path = _csv_path("Systems CMDB.xlsx")
        if xlsx_path.exists():
            try:
                from openpyxl import load_workbook

                wb = load_workbook(xlsx_path, read_only=True, data_only=True)
                ws = wb.worksheets[0]
                headers: list[str] = []
                for cell in ws[1]:
                    v = cell.value
                    if v is not None:
                        headers.append(str(v))
                wb.close()
                if headers:
                    return headers
            except Exception:
                pass
        # Fallback to merged CSV header
        csv_path = _csv_path("netbox_merged_export.csv")
        if csv_path.exists():
            with csv_path.open("r", encoding="utf-8") as fh:
                import csv as _csv

                reader = _csv.reader(fh)
                headers = next(reader, [])
                return [str(h) for h in headers if h]
    except Exception:
        pass
    return []


@router.get("/logs/tail")
def logs_tail(n: int = Query(200, ge=1, le=5000)) -> dict:
    """Return the last N lines of the export log.

    Response: { "lines": ["..", ".."] }
    """
    if not LOG_PATH.exists():
        return {"lines": []}
    try:
        with LOG_PATH.open("r", encoding="utf-8", errors="ignore") as fh:
            lines = fh.readlines()[-n:]
        return {"lines": [ln.rstrip("\n") for ln in lines]}
    except Exception:
        return {"lines": []}


@router.get("/config/ui")
def ui_config():
    """Get UI configuration (theme, etc.)."""
    theme = os.getenv("UI_THEME_DEFAULT", "nebula").strip() or "nebula"
    return {"theme_default": theme}


@router.get("/config/chat")
def chat_config():
    """Get chat configuration (system prompt, temperature)."""
    return {
        "system_prompt": os.getenv("CHAT_SYSTEM_PROMPT", ""),
        "temperature": _chat_default_temperature(),
    }

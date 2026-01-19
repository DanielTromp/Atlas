"""API endpoints for Puppet configuration and user management visualization."""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Annotated, Any

from fastapi import APIRouter, Body, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse

from infrastructure_atlas.application.services import create_puppet_service
from infrastructure_atlas.application.services.puppet import PuppetServiceProtocol
from infrastructure_atlas.infrastructure.external import GitClientError
from infrastructure_atlas.infrastructure.security.secret_store import SecretStoreUnavailable
from infrastructure_atlas.interfaces.api.dependencies import AdminUserDep, CurrentUserDep

router = APIRouter(prefix="/puppet", tags=["puppet"])


def get_puppet_service() -> PuppetServiceProtocol:
    """Return the Puppet service using the configured backend."""
    return create_puppet_service()


PuppetServiceDep = Annotated[PuppetServiceProtocol, Depends(get_puppet_service)]


class PuppetConfigCreate:
    """Create payload for Puppet configuration."""

    def __init__(
        self,
        name: str = Body(...),
        remote_url: str = Body(...),
        branch: str = Body("production"),
        ssh_key_path: str | None = Body(None),
    ):
        self.name = name
        self.remote_url = remote_url
        self.branch = branch
        self.ssh_key_path = ssh_key_path


class PuppetConfigUpdate:
    """Update payload for Puppet configuration."""

    def __init__(
        self,
        name: str | None = Body(None),
        remote_url: str | None = Body(None),
        branch: str | None = Body(None),
        ssh_key_path: str | None = Body(None),
    ):
        self.name = name
        self.remote_url = remote_url
        self.branch = branch
        self.ssh_key_path = ssh_key_path


CreateBody = Annotated[PuppetConfigCreate, Depends()]
UpdateBody = Annotated[PuppetConfigUpdate, Depends()]


def _meta_to_payload(meta: dict[str, object]) -> dict[str, object]:
    """Convert metadata to API payload format."""
    generated_at = meta.get("generated_at")
    if isinstance(generated_at, datetime):
        generated_at_str = generated_at.astimezone(UTC).isoformat().replace("+00:00", "Z")
    else:
        generated_at_str = None
    return {
        "generated_at": generated_at_str,
        "user_count": meta.get("user_count"),
        "group_count": meta.get("group_count"),
        "commit_hash": meta.get("commit_hash"),
        "commit_message": meta.get("commit_message"),
        "commit_date": meta.get("commit_date"),
        "source": meta.get("source"),
    }


@router.get("/configs")
def list_configs(admin: AdminUserDep, service: PuppetServiceDep):
    """List all Puppet configurations."""
    configs = service.list_configs()
    return [
        {
            "id": c.id,
            "name": c.name,
            "remote_url": c.remote_url,
            "branch": c.branch,
            "has_ssh_key": bool(c.ssh_key_secret),
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
        }
        for c in configs
    ]


@router.post("/configs")
def create_config(admin: AdminUserDep, payload: CreateBody, service: PuppetServiceDep):
    """Create a new Puppet configuration."""
    try:
        entity = service.create_config(
            name=payload.name,
            remote_url=payload.remote_url,
            branch=payload.branch,
            ssh_key_path=payload.ssh_key_path,
        )
    except SecretStoreUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return {
        "id": entity.id,
        "name": entity.name,
        "remote_url": entity.remote_url,
        "branch": entity.branch,
        "has_ssh_key": bool(entity.ssh_key_secret),
    }


@router.put("/configs/{config_id}")
def update_config(admin: AdminUserDep, config_id: str, payload: UpdateBody, service: PuppetServiceDep):
    """Update an existing Puppet configuration."""
    try:
        entity = service.update_config(
            config_id,
            name=payload.name,
            remote_url=payload.remote_url,
            branch=payload.branch,
            ssh_key_path=payload.ssh_key_path,
        )
    except SecretStoreUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except ValueError as exc:
        message = str(exc)
        if message.lower().endswith("not found"):
            raise HTTPException(status_code=404, detail=message) from exc
        raise HTTPException(status_code=400, detail=message) from exc

    return {
        "id": entity.id,
        "name": entity.name,
        "remote_url": entity.remote_url,
        "branch": entity.branch,
        "has_ssh_key": bool(entity.ssh_key_secret),
    }


@router.delete("/configs/{config_id}")
def delete_config(admin: AdminUserDep, config_id: str, service: PuppetServiceDep):
    """Delete a Puppet configuration."""
    try:
        removed = service.delete_config(config_id)
    except SecretStoreUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    if not removed:
        raise HTTPException(status_code=404, detail="Puppet configuration not found")
    return {"status": "deleted"}


@router.get("/instances")
def list_instances(user: CurrentUserDep, service: PuppetServiceDep):
    """List all Puppet instances for the web UI."""
    entries = service.list_configs_with_status()
    response: list[dict[str, object]] = []
    for config, meta in entries:
        meta_payload = _meta_to_payload(meta or {}) if meta else {
            "generated_at": None,
            "user_count": None,
            "group_count": None,
            "source": None,
        }
        response.append(
            {
                "id": config.id,
                "name": config.name,
                "remote_url": config.remote_url,
                "branch": config.branch,
                "has_ssh_key": bool(config.ssh_key_secret),
                "last_refresh": meta_payload.get("generated_at"),
                "user_count": meta_payload.get("user_count"),
                "group_count": meta_payload.get("group_count"),
                "commit_hash": meta_payload.get("commit_hash"),
            }
        )
    return response


@router.post("/{config_id}/refresh")
def refresh_inventory(
    config_id: str,
    request: Request,
    user: CurrentUserDep,
    service: PuppetServiceDep,
):
    """Refresh Puppet inventory from Git repository."""
    permissions = getattr(request.state, "permissions", frozenset())
    if "puppet.view" not in permissions and user.role != "admin":
        raise HTTPException(status_code=403, detail="Puppet access requires additional permissions")

    try:
        config, inventory, meta = service.refresh_inventory(config_id)
    except SecretStoreUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except ValueError as exc:
        message = str(exc)
        if message.lower().endswith("not found"):
            raise HTTPException(status_code=404, detail=message) from exc
        raise HTTPException(status_code=400, detail=message) from exc
    except GitClientError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    return {
        "config_id": config.id,
        "user_count": len(inventory.users),
        "group_count": len(inventory.groups),
        "meta": _meta_to_payload(meta or {}),
    }


@router.get("/users")
def list_users(
    user: CurrentUserDep,
    request: Request,
    service: PuppetServiceDep,
    config_id: str | None = None,
    search: str | None = None,
    refresh: bool = Query(False, description="Force refresh from Git repository"),
):
    """List all users from Puppet."""
    permissions = getattr(request.state, "permissions", frozenset())
    if "puppet.view" not in permissions and user.role != "admin":
        raise HTTPException(status_code=403, detail="Puppet access requires additional permissions")

    # Get config
    if config_id:
        config = service.get_config(config_id)
        if not config:
            raise HTTPException(status_code=404, detail="Puppet configuration not found")
    else:
        configs = service.list_configs()
        if not configs:
            raise HTTPException(status_code=404, detail="No Puppet configurations found")
        config = configs[0]

    try:
        _, inventory, meta = service.get_inventory(config.id, refresh=refresh)

        # Build user list with group access info
        users_list = []
        for username, puppet_user in inventory.users.items():
            # Get groups this user is a member of
            user_groups = [
                access.group_name
                for access in inventory.user_access
                if access.username == username
            ]

            # Apply search filter
            if search:
                search_lower = search.lower()
                if not (
                    search_lower in username.lower()
                    or (puppet_user.key_name and search_lower in puppet_user.key_name.lower())
                    or any(search_lower in g.lower() for g in user_groups)
                ):
                    continue

            users_list.append({
                "username": puppet_user.username,
                "uid": puppet_user.uid,
                "key_type": puppet_user.key_type,
                "key_name": puppet_user.key_name,
                "has_password": puppet_user.has_password,
                "has_ssh_key": puppet_user.has_ssh_key,
                "enabled": puppet_user.enabled,
                "has_sudo": username in inventory.sudo_users,
                "groups": user_groups,
                "is_removed": username in inventory.removed_users,
                # Security details
                "password_algorithm": puppet_user.password_algorithm,
                "account_locked": puppet_user.account_locked,
                "ssh_key_bits": puppet_user.ssh_key_bits,
            })

        # Sort by username
        users_list.sort(key=lambda u: u["username"].lower())

        return {
            "results": users_list,
            "total": len(users_list),
            "config_id": config.id,
            "meta": _meta_to_payload(meta or {}),
        }
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except GitClientError as exc:
        raise HTTPException(status_code=502, detail=f"Git error: {exc}") from exc


@router.get("/groups")
def list_groups(
    user: CurrentUserDep,
    request: Request,
    service: PuppetServiceDep,
    config_id: str | None = None,
    search: str | None = None,
    refresh: bool = Query(False, description="Force refresh from Git repository"),
):
    """List all groups from Puppet."""
    permissions = getattr(request.state, "permissions", frozenset())
    if "puppet.view" not in permissions and user.role != "admin":
        raise HTTPException(status_code=403, detail="Puppet access requires additional permissions")

    # Get config
    if config_id:
        config = service.get_config(config_id)
        if not config:
            raise HTTPException(status_code=404, detail="Puppet configuration not found")
    else:
        configs = service.list_configs()
        if not configs:
            raise HTTPException(status_code=404, detail="No Puppet configurations found")
        config = configs[0]

    try:
        _, inventory, meta = service.get_inventory(config.id, refresh=refresh)

        groups_list = []
        for group_name, puppet_group in inventory.groups.items():
            # Apply search filter
            if search:
                search_lower = search.lower()
                if not (
                    search_lower in group_name.lower()
                    or any(search_lower in m.lower() for m in puppet_group.members)
                ):
                    continue

            # Count members with sudo
            sudo_members = [m for m in puppet_group.members if m in inventory.sudo_users]

            groups_list.append({
                "name": puppet_group.name,
                "gid": puppet_group.gid,
                "members": puppet_group.members,
                "not_members": puppet_group.not_members,
                "member_count": len(puppet_group.members),
                "sudo_member_count": len(sudo_members),
            })

        # Sort by name
        groups_list.sort(key=lambda g: g["name"].lower())

        return {
            "results": groups_list,
            "total": len(groups_list),
            "config_id": config.id,
            "meta": _meta_to_payload(meta or {}),
        }
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except GitClientError as exc:
        raise HTTPException(status_code=502, detail=f"Git error: {exc}") from exc


@router.get("/users/{username}")
def get_user_detail(
    username: str,
    request: Request,
    user: CurrentUserDep,
    service: PuppetServiceDep,
    config_id: str | None = None,
):
    """Get detailed user information including all group memberships."""
    permissions = getattr(request.state, "permissions", frozenset())
    if "puppet.view" not in permissions and user.role != "admin":
        raise HTTPException(status_code=403, detail="Puppet access requires additional permissions")

    # Get config
    if config_id:
        config = service.get_config(config_id)
        if not config:
            raise HTTPException(status_code=404, detail="Puppet configuration not found")
    else:
        configs = service.list_configs()
        if not configs:
            raise HTTPException(status_code=404, detail="No Puppet configurations found")
        config = configs[0]

    try:
        _, inventory, meta = service.get_inventory(config.id, refresh=False)

        if username not in inventory.users:
            raise HTTPException(status_code=404, detail=f"User {username} not found")

        puppet_user = inventory.users[username]

        # Get all access entries for this user
        user_access = [
            {
                "group_name": access.group_name,
                "has_sudo": access.has_sudo,
                "access_type": access.access_type,
            }
            for access in inventory.user_access
            if access.username == username
        ]

        return {
            "username": puppet_user.username,
            "uid": puppet_user.uid,
            "key_type": puppet_user.key_type,
            "key_name": puppet_user.key_name,
            "has_password": puppet_user.has_password,
            "has_ssh_key": puppet_user.has_ssh_key,
            "enabled": puppet_user.enabled,
            "has_sudo": username in inventory.sudo_users,
            "is_removed": username in inventory.removed_users,
            "source_file": puppet_user.source_file,
            "access": user_access,
            "config_id": config.id,
            "meta": _meta_to_payload(meta or {}),
        }
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except GitClientError as exc:
        raise HTTPException(status_code=502, detail=f"Git error: {exc}") from exc


@router.get("/groups/{group_name}")
def get_group_detail(
    group_name: str,
    request: Request,
    user: CurrentUserDep,
    service: PuppetServiceDep,
    config_id: str | None = None,
):
    """Get detailed group information including all members."""
    permissions = getattr(request.state, "permissions", frozenset())
    if "puppet.view" not in permissions and user.role != "admin":
        raise HTTPException(status_code=403, detail="Puppet access requires additional permissions")

    # Get config
    if config_id:
        config = service.get_config(config_id)
        if not config:
            raise HTTPException(status_code=404, detail="Puppet configuration not found")
    else:
        configs = service.list_configs()
        if not configs:
            raise HTTPException(status_code=404, detail="No Puppet configurations found")
        config = configs[0]

    try:
        _, inventory, meta = service.get_inventory(config.id, refresh=False)

        if group_name not in inventory.groups:
            raise HTTPException(status_code=404, detail=f"Group {group_name} not found")

        puppet_group = inventory.groups[group_name]

        # Build member details
        members_detail = []
        for member in puppet_group.members:
            member_user = inventory.users.get(member)
            has_sudo = member in inventory.sudo_users
            members_detail.append({
                "username": member,
                "has_sudo": has_sudo,
                "enabled": member_user.enabled if member_user else False,
                "key_name": member_user.key_name if member_user else None,
            })

        return {
            "name": puppet_group.name,
            "gid": puppet_group.gid,
            "members": members_detail,
            "not_members": puppet_group.not_members,
            "source_file": puppet_group.source_file,
            "config_id": config.id,
            "meta": _meta_to_payload(meta or {}),
        }
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except GitClientError as exc:
        raise HTTPException(status_code=502, detail=f"Git error: {exc}") from exc


@router.get("/access-matrix")
def get_access_matrix(
    user: CurrentUserDep,
    request: Request,
    service: PuppetServiceDep,
    config_id: str | None = None,
):
    """Get a matrix view of user-group access rights.

    Returns a structure suitable for building a user/group access matrix visualization.
    """
    permissions = getattr(request.state, "permissions", frozenset())
    if "puppet.view" not in permissions and user.role != "admin":
        raise HTTPException(status_code=403, detail="Puppet access requires additional permissions")

    # Get config
    if config_id:
        config = service.get_config(config_id)
        if not config:
            raise HTTPException(status_code=404, detail="Puppet configuration not found")
    else:
        configs = service.list_configs()
        if not configs:
            raise HTTPException(status_code=404, detail="No Puppet configurations found")
        config = configs[0]

    try:
        _, inventory, meta = service.get_inventory(config.id, refresh=False)

        # Build access matrix
        # Structure: { username: { group_name: { has_sudo, access_type } } }
        matrix: dict[str, dict[str, Any]] = {}

        for access in inventory.user_access:
            if access.username not in matrix:
                user_obj = inventory.users.get(access.username)
                matrix[access.username] = {
                    "_user": {
                        "enabled": user_obj.enabled if user_obj else False,
                        "has_sudo_any": access.username in inventory.sudo_users,
                        "key_name": user_obj.key_name if user_obj else None,
                    }
                }
            matrix[access.username][access.group_name] = {
                "has_sudo": access.has_sudo,
                "access_type": access.access_type,
            }

        # Get sorted lists
        all_users = sorted(matrix.keys(), key=str.lower)
        all_groups = sorted(inventory.groups.keys(), key=str.lower)

        return {
            "users": all_users,
            "groups": all_groups,
            "matrix": matrix,
            "config_id": config.id,
            "meta": _meta_to_payload(meta or {}),
        }
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except GitClientError as exc:
        raise HTTPException(status_code=502, detail=f"Git error: {exc}") from exc


@router.get("/export")
def export_puppet_data(
    user: CurrentUserDep,
    request: Request,
    service: PuppetServiceDep,
    config_id: str | None = None,
):
    """Export Puppet data as Excel file with Users, Groups, and Access Matrix sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    permissions = getattr(request.state, "permissions", frozenset())
    if "puppet.view" not in permissions and user.role != "admin":
        raise HTTPException(status_code=403, detail="Puppet access requires additional permissions")

    # Get config
    if config_id:
        config = service.get_config(config_id)
        if not config:
            raise HTTPException(status_code=404, detail="Puppet configuration not found")
    else:
        configs = service.list_configs()
        if not configs:
            raise HTTPException(status_code=404, detail="No Puppet configurations found")
        config = configs[0]

    try:
        _, inventory, meta = service.get_inventory(config.id, refresh=False)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except GitClientError as exc:
        raise HTTPException(status_code=502, detail=f"Git error: {exc}") from exc

    # Create workbook
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    weak_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    weak_font = Font(color="DC2626")
    locked_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    sudo_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    member_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    # ===== Sheet 1: Users =====
    ws_users = wb.active
    ws_users.title = "Users"

    user_headers = [
        "Username", "UID", "Email", "Status", "Sudo Access", "Groups",
        "Password Algorithm", "SSH Key Type", "SSH Key Bits", "Security Notes"
    ]
    for col, header in enumerate(user_headers, 1):
        cell = ws_users.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Sort users by username
    sorted_users = sorted(inventory.users.values(), key=lambda u: u.username.lower())

    for row_idx, puppet_user in enumerate(sorted_users, 2):
        username = puppet_user.username

        # Determine status
        if puppet_user.account_locked:
            status = "Locked"
        elif username in inventory.removed_users:
            status = "Removed"
        elif not puppet_user.enabled:
            status = "Disabled"
        else:
            status = "Active"

        # Get groups
        user_groups = [
            access.group_name
            for access in inventory.user_access
            if access.username == username
        ]

        # Security notes
        notes = []
        if puppet_user.password_algorithm == "md5":
            notes.append("⚠️ MD5 hash is weak - user should change password")
        if puppet_user.ssh_key_bits and puppet_user.ssh_key_bits < 2048:
            notes.append(f"⚠️ SSH key too short ({puppet_user.ssh_key_bits} bits)")

        row_data = [
            puppet_user.username,
            puppet_user.uid,
            puppet_user.key_name or "",
            status,
            "Yes" if username in inventory.sudo_users else "No",
            ", ".join(user_groups),
            (puppet_user.password_algorithm or "").upper() if puppet_user.has_password else "None",
            puppet_user.key_type or "" if puppet_user.has_ssh_key else "None",
            puppet_user.ssh_key_bits if puppet_user.has_ssh_key else "",
            "; ".join(notes),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws_users.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

            # Highlight weak security
            if col == 7 and puppet_user.password_algorithm == "md5":  # Password Algorithm
                cell.fill = weak_fill
                cell.font = weak_font
            if col == 9 and puppet_user.ssh_key_bits and puppet_user.ssh_key_bits < 2048:  # SSH Key Bits
                cell.fill = weak_fill
                cell.font = weak_font
            if col == 4 and status == "Locked":  # Status
                cell.fill = locked_fill
            if col == 5 and username in inventory.sudo_users:  # Sudo
                cell.fill = sudo_fill

    # Auto-width columns
    for col in range(1, len(user_headers) + 1):
        ws_users.column_dimensions[get_column_letter(col)].width = 18
    ws_users.column_dimensions["F"].width = 40  # Groups
    ws_users.column_dimensions["J"].width = 50  # Security Notes

    # Freeze header row
    ws_users.freeze_panes = "A2"

    # ===== Sheet 2: Groups =====
    ws_groups = wb.create_sheet("Groups")

    group_headers = ["Group Name", "GID", "Member Count", "Members", "Removed Members"]
    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    sorted_groups = sorted(inventory.groups.values(), key=lambda g: g.name.lower())

    for row_idx, group in enumerate(sorted_groups, 2):
        row_data = [
            group.name,
            group.gid,
            len(group.members),
            ", ".join(sorted(group.members)),
            ", ".join(sorted(group.not_members)) if group.not_members else "",
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws_groups.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    for col in range(1, len(group_headers) + 1):
        ws_groups.column_dimensions[get_column_letter(col)].width = 18
    ws_groups.column_dimensions["D"].width = 60  # Members
    ws_groups.column_dimensions["E"].width = 30  # Removed Members
    ws_groups.freeze_panes = "A2"

    # ===== Sheet 3: Access Matrix =====
    ws_matrix = wb.create_sheet("Access Matrix")

    all_groups = sorted(inventory.groups.keys(), key=str.lower)
    all_usernames = sorted(inventory.users.keys(), key=str.lower)

    # Build matrix data
    matrix_headers = ["Username", "Sudo Access"] + all_groups
    for col, header in enumerate(matrix_headers, 1):
        cell = ws_matrix.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Build user->group membership lookup
    user_group_map: dict[str, set[str]] = {u: set() for u in all_usernames}
    for access in inventory.user_access:
        if access.username in user_group_map:
            user_group_map[access.username].add(access.group_name)

    for row_idx, username in enumerate(all_usernames, 2):
        has_sudo = username in inventory.sudo_users
        user_groups = user_group_map.get(username, set())

        # Username cell
        cell = ws_matrix.cell(row=row_idx, column=1, value=username)
        cell.border = thin_border

        # Sudo cell
        cell = ws_matrix.cell(row=row_idx, column=2, value="Yes" if has_sudo else "")
        cell.border = thin_border
        if has_sudo:
            cell.fill = sudo_fill

        # Group membership cells
        for col_idx, group_name in enumerate(all_groups, 3):
            is_member = group_name in user_groups
            cell = ws_matrix.cell(row=row_idx, column=col_idx, value="✓" if is_member else "")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if is_member:
                cell.fill = member_fill

    # Set column widths for matrix
    ws_matrix.column_dimensions["A"].width = 18
    ws_matrix.column_dimensions["B"].width = 12
    for col_idx in range(3, len(all_groups) + 3):
        ws_matrix.column_dimensions[get_column_letter(col_idx)].width = 14
    ws_matrix.freeze_panes = "C2"

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generate filename
    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    filename = f"puppet_{config.name.replace(' ', '_')}_{timestamp}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


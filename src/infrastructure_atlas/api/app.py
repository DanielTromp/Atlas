from __future__ import annotations

import asyncio
import csv
import html
import json
import math
import os
import re
import secrets
import shlex
import shutil
import sys
import time
import uuid
import warnings
from collections import Counter
from collections.abc import Callable, Collection, Iterable, Mapping, Sequence
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import UTC, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from threading import Lock
from typing import Annotated, Any, Literal
from zoneinfo import ZoneInfo

import duckdb
import numpy as np
import pandas as pd
import requests
from dotenv import dotenv_values
from fastapi import Body, FastAPI, HTTPException, Query, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import (
    FileResponse,
    HTMLResponse,
    JSONResponse,
    PlainTextResponse,
    RedirectResponse,
    StreamingResponse,
)
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.sessions import SessionMiddleware

from infrastructure_atlas import backup_manager, backup_sync

# AI/LangChain functionality disabled
# from infrastructure_atlas.application.chat_agents import AgentRuntime, AgentRuntimeError
from infrastructure_atlas.application.dto import (
    AdminEnvResponseDTO,
    BackupInfoDTO,
    EnvSettingDTO,
    SuggestionCommentDTO,
    SuggestionCommentResponseDTO,
    SuggestionItemDTO,
    SuggestionListDTO,
    meta_to_dto,
    suggestion_to_dto,
    suggestions_to_dto,
)
from infrastructure_atlas.application.role_defaults import DEFAULT_ROLE_DEFINITIONS
from infrastructure_atlas.application.security import hash_password, verify_password
from infrastructure_atlas.application.services import create_vcenter_service
from infrastructure_atlas.application.services.suggestions import (
    SUGGESTION_CLASSIFICATIONS,
    SUGGESTION_STATUSES,
    SuggestionNotFoundError,
    create_suggestion_service,
    get_suggestion_backend_type,
    migrate_csv_to_airtable,
)
from infrastructure_atlas.db import get_sessionmaker
from infrastructure_atlas.db.setup import init_database
from infrastructure_atlas.db.models import (
    ChatMessage,
    ChatSession,
    GlobalAPIKey,
    RolePermission,
    User,
    UserAPIKey,
)
from infrastructure_atlas.domain.integrations.commvault import (
    CommvaultJob,
    CommvaultJobList,
    CommvaultPlan,
    CommvaultStoragePool,
)
from infrastructure_atlas.env import load_env, project_root
from infrastructure_atlas.infrastructure.external import (
    ZabbixAuthError,
    ZabbixClient,
    ZabbixClientConfig,
    ZabbixConfigError,
    ZabbixError,
)
from infrastructure_atlas.infrastructure.external.commvault_client import (
    CommvaultClient,
    CommvaultClientConfig,
    CommvaultError,
)
from infrastructure_atlas.infrastructure.logging import get_logger, logging_context, setup_logging
from infrastructure_atlas.infrastructure.metrics import get_metrics_snapshot, snapshot_to_prometheus
from infrastructure_atlas.infrastructure.tracing import init_tracing, tracing_enabled
from infrastructure_atlas.interfaces.api import bootstrap_api
from infrastructure_atlas.interfaces.api.dependencies import (
    CurrentUserDep,
    DbSessionDep,
    OptionalUserDep,
)
from infrastructure_atlas.interfaces.api.middleware import ObservabilityMiddleware
from infrastructure_atlas.interfaces.api.routes import tools as tools_router
from infrastructure_atlas.interfaces.api.routes.confluence import confluence_search
from infrastructure_atlas.interfaces.api.routes.jira import jira_search
from infrastructure_atlas.interfaces.api.routes.netbox import netbox_search
from infrastructure_atlas.interfaces.api.schemas import ToolDefinition

try:
    from openai import OpenAI  # type: ignore
except Exception:  # optional dependency
    OpenAI = None  # type: ignore

try:  # optional dependency; urllib3 may not always be available
    from urllib3 import disable_warnings as _disable_urllib3_warnings
    from urllib3.exceptions import InsecureRequestWarning as _InsecureRequestWarning
except Exception:  # pragma: no cover - urllib3 optional
    _disable_urllib3_warnings = None  # type: ignore[assignment]
    _InsecureRequestWarning = None  # type: ignore[assignment]

load_env()

AMS_TZ = ZoneInfo("Europe/Amsterdam")

logger = get_logger(__name__)
if tracing_enabled():
    init_tracing("infrastructure-atlas-api")

# Ensure the application database is migrated to the latest revision so
# authentication state can be loaded lazily by request handlers.
try:
    init_database()
except Exception as exc:  # pragma: no cover - surfaces during boot
    logger.exception("Failed to initialise database during API startup", extra={"error": str(exc)})
    raise
else:
    # Alembic can override logging handlers; ensure our configuration is still active.
    setup_logging()
    logger.disabled = False

SessionLocal = get_sessionmaker()


class _TaskLogger:
    """Mutable container passed into task_logging for success metadata."""

    __slots__ = ("_success_extra",)

    def __init__(self) -> None:
        self._success_extra: dict[str, Any] = {}

    def add_success(self, **kwargs: Any) -> None:
        for key, value in kwargs.items():
            if value is not None:
                self._success_extra[key] = value

    @property
    def success_extra(self) -> dict[str, Any]:
        return self._success_extra


@contextmanager
def task_logging(task: str, **context: Any):
    """Log lifecycle events around long-running UI-triggered tasks."""

    start = time.perf_counter()
    tracker = _TaskLogger()
    safe_context = {k: v for k, v in context.items() if v not in (None, "")}

    with logging_context(task=task, **safe_context):
        logger.info("Task started", extra={"event": "task_started"})
        try:
            yield tracker
        except Exception:
            duration_ms = int((time.perf_counter() - start) * 1000)
            with logging_context(duration_ms=duration_ms, **tracker.success_extra):
                logger.exception("Task failed", extra={"event": "task_failed"})
            raise
        else:
            duration_ms = int((time.perf_counter() - start) * 1000)
            with logging_context(duration_ms=duration_ms, **tracker.success_extra):
                logger.info("Task completed", extra={"event": "task_completed"})


_USAGE_FIELD_ALIASES: dict[str, str] = {
    "prompt_tokens": "prompt_tokens",
    "completion_tokens": "completion_tokens",
    "total_tokens": "total_tokens",
    "input_tokens": "prompt_tokens",
    "output_tokens": "completion_tokens",
    "usage_tokens": "total_tokens",
    "input_token_count": "prompt_tokens",
    "output_token_count": "completion_tokens",
    "total_token_count": "total_tokens",
    "promptTokenCount": "prompt_tokens",
    "candidatesTokenCount": "completion_tokens",
    "totalTokens": "total_tokens",
}


def _normalise_usage(raw: Any) -> dict[str, int] | None:
    if raw is None:
        return None
    usage: dict[str, int] = {}
    for source, target in _USAGE_FIELD_ALIASES.items():
        value = None
        if isinstance(raw, dict):
            value = raw.get(source)
        else:
            value = getattr(raw, source, None)
        if isinstance(value, int | float):
            usage[target] = int(value)
    return usage or None


def _safe_json_loads(data: str) -> Any | None:
    try:
        return json.loads(data)
    except Exception as exc:  # pragma: no cover - defensive logging
        logger.debug("Skipping streaming event with invalid JSON", extra={"error": str(exc)})
        return None


def require_permission(request: Request, permission: str) -> None:
    user = getattr(request.state, "user", None)
    if user is None:
        return
    if getattr(user, "role", "") == "admin":
        return
    permissions = getattr(request.state, "permissions", frozenset())
    if permission not in permissions:
        raise HTTPException(status_code=403, detail="Forbidden: missing permission")


def _env_flag(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


METRICS_ENABLED = _env_flag("ATLAS_METRICS_ENABLED")
METRICS_TOKEN = os.getenv("ATLAS_METRICS_TOKEN", "").strip()
METRICS_MEDIA_TYPE = "text/plain; version=0.0.4; charset=utf-8"


def ensure_default_role_permissions() -> None:
    with SessionLocal() as db:
        existing = {record.role: record for record in db.execute(select(RolePermission)).scalars()}
        changed = False
        for role, spec in DEFAULT_ROLE_DEFINITIONS.items():
            label = (spec.get("label") or role).strip() or role
            description = spec.get("description") or None
            default_perms = sorted({str(p).strip() for p in spec.get("permissions", []) if str(p).strip()})
            record = existing.get(role)
            if record is None:
                record = RolePermission(
                    role=role,
                    label=label,
                    description=description,
                    permissions=default_perms,
                )
                db.add(record)
                changed = True
            else:
                updated = False
                if not record.label:
                    record.label = label
                    updated = True
                if description and not record.description:
                    record.description = description
                    updated = True
                current = set(record.permissions or [])
                desired = current.union(default_perms)
                if desired != current:
                    record.permissions = sorted(desired)
                    updated = True
                if updated:
                    db.add(record)
                    changed = True
        if changed:
            db.commit()


def ensure_default_admin() -> None:
    with SessionLocal() as db:
        existing = db.execute(select(User.id).limit(1)).scalar_one_or_none()
        if existing:
            return

        username = os.getenv("ATLAS_DEFAULT_ADMIN_USERNAME", "admin").strip().lower() or "admin"
        seed_password = os.getenv("ATLAS_DEFAULT_ADMIN_PASSWORD", "").strip() or UI_PASSWORD

        if not seed_password:
            logger.warning(
                "No users exist and ATLAS_DEFAULT_ADMIN_PASSWORD is not set; set it (or ATLAS_UI_PASSWORD) to bootstrap the first login.",
                extra={"username": username},
            )
            return

        user = User(
            username=username,
            display_name="Administrator",
            role="admin",
            password_hash=hash_password(seed_password),
            is_active=True,
        )
        db.add(user)
        db.commit()
        logger.info(
            "Created default admin user",
            extra={"username": username},
        )


ensure_default_role_permissions()
ensure_default_admin()

# Metrics configuration
METRICS_ENABLED = _env_flag("ATLAS_METRICS_ENABLED")
METRICS_TOKEN = os.getenv("ATLAS_METRICS_TOKEN", "").strip()
METRICS_MEDIA_TYPE = "text/plain; version=0.0.4; charset=utf-8"

# Auth configuration
API_TOKEN = os.getenv("ATLAS_API_TOKEN", "").strip()
UI_PASSWORD = os.getenv("ATLAS_UI_PASSWORD", "").strip()  # legacy fallback
UI_SECRET = os.getenv("ATLAS_UI_SECRET", "").strip() or secrets.token_hex(32)
SESSION_COOKIE_NAME = "atlas_ui"
SESSION_USER_KEY = "user_id"

# Initialize FastAPI application
app = FastAPI(title="Infrastructure Atlas API", version="0.1.0")
app.add_middleware(
    ObservabilityMiddleware,
    metrics_enabled=METRICS_ENABLED,
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)


def _require_metrics_token(request: Request) -> None:
    if not METRICS_TOKEN:
        return
    auth = request.headers.get("Authorization", "")
    if not auth.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Unauthorized")
    token = auth[7:].strip()
    if not secrets.compare_digest(token, METRICS_TOKEN):
        raise HTTPException(status_code=401, detail="Unauthorized")


if METRICS_ENABLED:

    @app.get("/metrics", response_class=PlainTextResponse)
    def metrics_endpoint(request: Request) -> PlainTextResponse:
        _require_metrics_token(request)
        payload = snapshot_to_prometheus(get_metrics_snapshot())
        return PlainTextResponse(payload, media_type=METRICS_MEDIA_TYPE)


# Auth middleware and helpers
class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path or "/"

        request.state.user = None
        request.state.permissions = frozenset()
        user_id = request.session.get(SESSION_USER_KEY) if hasattr(request, "session") else None
        if user_id:
            with SessionLocal() as db:
                user = db.get(User, user_id)
                if user and user.is_active:
                    perms_record = db.get(RolePermission, user.role)
                    perms = frozenset((perms_record.permissions or []) if perms_record else [])
                    request.state.user = user
                    request.state.permissions = perms
                    try:
                        user.permissions = perms
                    except Exception:
                        pass
                else:
                    request.session.pop(SESSION_USER_KEY, None)
                    request.state.user = None
                    request.state.permissions = frozenset()

        # Public endpoints
        if path in ("/favicon.ico", "/favicon.png", "/health"):
            return await call_next(request)

        if path == "/":
            if not _has_ui_session(request):
                return RedirectResponse(url="/auth/login")
            return await call_next(request)
        if path.startswith("/app"):
            if not _has_ui_session(request):
                next_url = path
                if request.url.query:
                    next_url += f"?{request.url.query}"
                return RedirectResponse(url=f"/auth/login?next={next_url}")
            return await call_next(request)

        # Auth endpoints are always allowed
        if path.startswith("/auth/"):
            return await call_next(request)

        # API protection via Bearer token; allow UI session as alternative
        if API_TOKEN and _is_api_path(path):
            if _has_bearer_token(request) or _has_ui_session(request):
                return await call_next(request)
            # Unauthorized
            return JSONResponse({"detail": "Unauthorized"}, status_code=401, headers={"WWW-Authenticate": "Bearer"})

        return await call_next(request)


def _has_bearer_token(request: Request) -> bool:
    if not API_TOKEN:
        return False
    auth = request.headers.get("Authorization", "")
    if not auth.lower().startswith("bearer "):
        return False
    token = auth[7:].strip()
    return secrets.compare_digest(token, API_TOKEN)


def _has_ui_session(request: Request) -> bool:
    try:
        return bool(request.session.get(SESSION_USER_KEY))
    except Exception:
        return False


def _is_api_path(path: str) -> bool:
    # Treat everything except static/frontend and auth endpoints as API
    if path.startswith("/app") or path.startswith("/auth"):
        return False
    if path in ("/", "/favicon.ico"):
        return False
    if path == "/metrics":
        return False
    return True


# Reduce aggressive caching for the static UI during development to avoid stale assets
@app.middleware("http")
async def no_cache_static(request, call_next):
    response = await call_next(request)
    try:
        path = request.url.path
        if path.startswith("/app"):
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
    except Exception:
        pass
    return response


def _data_dir() -> Path:
    root = project_root()
    raw = os.getenv("NETBOX_DATA_DIR", "data")
    path = Path(raw) if os.path.isabs(raw) else (root / raw)
    path.mkdir(parents=True, exist_ok=True)
    return path


def _csv_path(name: str) -> Path:
    return _data_dir() / name


# Simple export log file (appends)
LOG_PATH = project_root() / "export.log"

# Dataset task definitions and metadata
CommandBuilder = Callable[["DatasetDefinition", "DatasetMetadata"], list[str] | None]


@dataclass(frozen=True, slots=True)
class DatasetDefinition:
    id: str
    label: str
    path_globs: tuple[str, ...]
    description: str | None = None
    since_source: str | None = None
    context: dict[str, Any] = field(default_factory=dict)
    command_builder: CommandBuilder | None = None


@dataclass(slots=True)
class DatasetFileRecord:
    relative_path: str
    absolute_path: Path
    exists: bool
    size_bytes: int | None
    modified: datetime | None


@dataclass(slots=True)
class DatasetMetadata:
    definition: DatasetDefinition
    files: list[DatasetFileRecord]
    last_updated: datetime | None
    extras: dict[str, Any] = field(default_factory=dict)

    def find_file(self, relative_path: str) -> DatasetFileRecord | None:
        for record in self.files:
            if record.relative_path == relative_path:
                return record
        return None


def _command_with_uv(args: Sequence[str]) -> list[str]:
    if shutil.which("uv"):
        return ["uv", "run", *args]
    return [sys.executable, "-m", "infrastructure_atlas.cli", *args]


def _build_netbox_cache_command(defn: DatasetDefinition, meta: DatasetMetadata) -> list[str]:
    return _command_with_uv(["atlas", "export", "cache"])


def _make_vcenter_command_builder(config_id: str) -> CommandBuilder:
    def _builder(_: DatasetDefinition, __: DatasetMetadata) -> list[str]:
        return _command_with_uv(["atlas", "vcenter", "refresh", "--id", config_id])

    return _builder


def _collect_task_dataset_definitions() -> list[DatasetDefinition]:
    definitions: list[DatasetDefinition] = [
        DatasetDefinition(
            id="netbox-cache",
            label="NetBox Cache",
            path_globs=("netbox_cache.json",),
            description="Cached NetBox snapshot used for export diffing.",
            command_builder=_build_netbox_cache_command,
        ),
    ]
    definitions.extend(_discover_vcenter_task_definitions())
    return definitions


def _discover_vcenter_task_definitions() -> list[DatasetDefinition]:
    definitions: list[DatasetDefinition] = []
    known_configs: dict[str, str | None] = {}
    try:
        with SessionLocal() as session:
            service = create_vcenter_service(session)
            entries = service.list_configs_with_status()
    except Exception:
        entries = []
    for config, _meta in entries:
        label = f"vCenter: {config.name or config.id}"
        definitions.append(
            DatasetDefinition(
                id=f"vcenter-{config.id}",
                label=label,
                path_globs=(f"vcenter/{config.id}.json",),
                description="Cached vCenter inventory snapshot.",
                since_source=None,
                context={
                    "config_id": config.id,
                    "config_name": config.name,
                },
                command_builder=_make_vcenter_command_builder(config.id),
            )
        )
        known_configs[config.id] = config.name

    vcenter_dir = _data_dir() / "vcenter"
    if vcenter_dir.exists():
        for path in sorted(vcenter_dir.glob("*.json")):
            config_id = path.stem
            if config_id in known_configs:
                continue
            label = f"vCenter cache ({config_id})"
            definitions.append(
                DatasetDefinition(
                    id=f"vcenter-{config_id}",
                    label=label,
                    path_globs=(f"vcenter/{path.name}",),
                    description="Cached vCenter inventory snapshot.",
                    context={
                        "config_id": config_id,
                        "config_name": None,
                        "orphan": True,
                    },
                    command_builder=_make_vcenter_command_builder(config_id),
                )
            )
    return definitions


def _dataset_file_record(base: Path, path: Path) -> DatasetFileRecord:
    try:
        relative = path.relative_to(base)
        relative_str = relative.as_posix()
    except ValueError:
        relative_str = path.as_posix()
    exists = path.exists()
    size = None
    modified: datetime | None = None
    if exists:
        try:
            stat = path.stat()
        except FileNotFoundError:
            exists = False
        else:
            size = stat.st_size
            modified = datetime.fromtimestamp(stat.st_mtime, tz=UTC)
    return DatasetFileRecord(
        relative_path=relative_str,
        absolute_path=path,
        exists=exists,
        size_bytes=size,
        modified=modified,
    )


def _build_dataset_metadata(defn: DatasetDefinition) -> DatasetMetadata:
    base = _data_dir()
    files: list[DatasetFileRecord] = []
    for pattern in defn.path_globs:
        pattern_norm = (pattern or "").strip()
        if not pattern_norm:
            continue
        if any(ch in pattern_norm for ch in "*?["):
            matches = sorted(base.glob(pattern_norm))
            for match in matches:
                files.append(_dataset_file_record(base, match))
        else:
            files.append(_dataset_file_record(base, base / pattern_norm))
    last_updated: datetime | None = None
    for record in files:
        if record.modified and (last_updated is None or record.modified > last_updated):
            last_updated = record.modified
    return DatasetMetadata(definition=defn, files=files, last_updated=last_updated)


def _build_dataset_command(defn: DatasetDefinition, meta: DatasetMetadata) -> list[str] | None:
    if defn.command_builder is None:
        return None
    return defn.command_builder(defn, meta)


# ---------------------------
# Commvault helper functions
# ---------------------------

COMMVAULT_BACKUPS_JSON = "commvault_backups.json"
COMMVAULT_STORAGE_JSON = "commvault_storage.json"
COMMVAULT_PLANS_JSON = "commvault_plans.json"


def _commvault_client_from_env() -> CommvaultClient:
    base_url = os.getenv("COMMVAULT_BASE_URL", "").strip()
    if not base_url:
        raise RuntimeError("Commvault integration is not configured (COMMVAULT_BASE_URL missing)")

    authtoken = os.getenv("COMMVAULT_API_TOKEN", "").strip() or None
    username = os.getenv("COMMVAULT_EMAIL", "").strip() or None
    password = os.getenv("COMMVAULT_PASSWORD", "")
    if password:
        password = password.strip()
    if not authtoken and not (username and password):
        raise RuntimeError("Commvault credentials are not configured (token or email/password required)")

    verify_tls = _env_flag("COMMVAULT_VERIFY_TLS", True)
    timeout_raw = os.getenv("COMMVAULT_TIMEOUT", "").strip()
    try:
        timeout = float(timeout_raw) if timeout_raw else 30.0
    except ValueError:
        timeout = 30.0

    if not verify_tls:
        _maybe_disable_commvault_tls_warnings()

    config = CommvaultClientConfig(
        base_url=base_url,
        authtoken=authtoken,
        username=username,
        password=password or None,
        verify_tls=verify_tls,
        timeout=timeout,
    )
    return CommvaultClient(config)


def _maybe_disable_commvault_tls_warnings() -> None:
    """Mute urllib3 TLS warnings when verification is intentionally disabled."""

    if _disable_urllib3_warnings and _InsecureRequestWarning:
        _disable_urllib3_warnings(_InsecureRequestWarning)
        return
    warnings.filterwarnings("ignore", category=Warning, module="urllib3")


def _collect_commvault_jobs_for_ui(
    client: CommvaultClient,
    *,
    limit: int,
    offset: int,
    since: datetime | None,
    tracked_job_ids: Collection[int] | None = None,
    latest_start: datetime | None = None,
    cached_recent_count: int | None = None,
    cutoff_start: datetime | None = None,
) -> CommvaultJobList:
    if limit < 0:
        raise ValueError("limit must be zero or positive")

    job_type = None
    page_target = 200 if limit == 0 else max(1, min(500, limit))
    tracked_remaining: set[int] = {int(job_id) for job_id in (tracked_job_ids or [])}
    threshold: datetime | None = since
    if latest_start and (threshold is None or latest_start < threshold):
        threshold = latest_start
    if cutoff_start and (threshold is None or cutoff_start < threshold):
        threshold = cutoff_start

    def _list_jobs_with_retry(*, fetch_limit: int, fetch_offset: int) -> CommvaultJobList:
        attempt_limit = max(1, fetch_limit)
        last_error: CommvaultError | None = None
        while attempt_limit >= 1:
            try:
                return client.list_jobs(
                    limit=attempt_limit,
                    offset=fetch_offset,
                    job_type=job_type,
                )
            except CommvaultError as exc:
                last_error = exc
                if attempt_limit <= 25:
                    break
                attempt_limit = max(25, attempt_limit // 2)
        if last_error is not None:
            raise last_error
        return CommvaultJobList(total_available=0, jobs=())

    try:
        initial = _list_jobs_with_retry(fetch_limit=1, fetch_offset=0)
    except CommvaultError:
        initial = _list_jobs_with_retry(fetch_limit=25, fetch_offset=0)

    total_available = initial.total_available or len(initial.jobs)
    if not total_available:
        return CommvaultJobList(total_available=0, jobs=())

    target_end = max(0, total_available - offset)
    if target_end <= 0:
        return CommvaultJobList(total_available=total_available, jobs=())

    approximate_needed = page_target
    if cached_recent_count is not None:
        approximate_needed = max(approximate_needed, min(total_available, cached_recent_count + page_target))
    if tracked_remaining:
        approximate_needed = max(approximate_needed, min(total_available, len(tracked_remaining) + page_target))
    if limit > 0:
        approximate_needed = max(approximate_needed, min(total_available, limit + page_target))

    target_start = max(0, target_end - approximate_needed)
    jobs: list[CommvaultJob] = []
    seen: set[int] = set()
    current_end = target_end

    while current_end > target_start and (limit == 0 or len(jobs) < limit):
        request_limit = min(page_target, current_end - target_start)
        try:
            response = _list_jobs_with_retry(
                fetch_limit=request_limit,
                fetch_offset=current_end - request_limit,
            )
        except CommvaultError:
            break

        batch = list(response.jobs)
        if not batch:
            break

        for job in reversed(batch):
            try:
                job_id = int(job.job_id)
            except (TypeError, ValueError):
                job_id = None
            if job_id is not None and job_id in seen:
                continue
            jobs.append(job)
            if job_id is not None:
                seen.add(job_id)
                tracked_remaining.discard(job_id)

        current_end -= len(batch)

        if limit > 0 and len(jobs) >= limit:
            break
        if tracked_remaining:
            continue
        if threshold:
            oldest = batch[0].start_time if batch else None
            if oldest and oldest < threshold:
                break
        if len(batch) < request_limit:
            break

    if limit > 0 and len(jobs) > limit:
        jobs = jobs[:limit]
    return CommvaultJobList(total_available=total_available, jobs=tuple(jobs))


def _serialise_commvault_job(job: CommvaultJob) -> dict[str, Any]:
    def iso(dt: datetime | None) -> str | None:
        if not dt:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        return dt.astimezone(UTC).isoformat()

    return {
        "job_id": job.job_id,
        "job_type": job.job_type,
        "status": job.status,
        "localized_status": job.localized_status,
        "localized_operation": job.localized_operation,
        "client_name": job.client_name,
        "client_id": job.client_id,
        "destination_client_name": job.destination_client_name,
        "subclient_name": job.subclient_name,
        "backup_set_name": job.backup_set_name,
        "application_name": job.application_name,
        "backup_level_name": job.backup_level_name,
        "plan_name": job.plan_name,
        "client_groups": list(job.client_groups),
        "storage_policy_name": job.storage_policy_name,
        "start_time": iso(job.start_time),
        "end_time": iso(job.end_time),
        "elapsed_seconds": job.elapsed_seconds,
        "size_of_application_bytes": job.size_of_application_bytes,
        "size_on_media_bytes": job.size_on_media_bytes,
        "total_num_files": job.total_num_files,
        "percent_complete": job.percent_complete,
        "percent_savings": job.percent_savings,
        "average_throughput_gb_per_hr": job.average_throughput,
        "retain_until": iso(job.retain_until),
    }


def _load_commvault_backups() -> dict[str, Any]:
    path = _data_dir() / COMMVAULT_BACKUPS_JSON
    if not path.exists():
        return {
            "jobs": [],
            "generated_at": None,
            "total_cached": 0,
            "version": 2,
        }
    try:
        with path.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
            if isinstance(data, Mapping):
                jobs = list(data.get("jobs") or [])
                generated = data.get("generated_at")
                total_cached = data.get("total_cached")
                try:
                    total_cached = int(total_cached)
                except (TypeError, ValueError):
                    total_cached = len(jobs)
                return {
                    "jobs": jobs,
                    "generated_at": generated,
                    "total_cached": total_cached,
                    "version": data.get("version", 2),
                }
    except Exception:
        logger.exception("Failed to load cached Commvault backups", extra={"event": "commvault_cache_error"})
    return {
        "jobs": [],
        "generated_at": None,
        "total_cached": 0,
        "version": 2,
    }


def _parse_job_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=UTC)
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return parsed


# Suggestion service instance (lazy-loaded)
_suggestion_service = None
_suggestion_service_lock = Lock()


def _get_suggestion_service():
    """Get or create the suggestion service instance."""
    global _suggestion_service
    if _suggestion_service is None:
        with _suggestion_service_lock:
            if _suggestion_service is None:
                _suggestion_service = create_suggestion_service(_data_dir())
    return _suggestion_service


def _reset_suggestion_service():
    """Reset the suggestion service (for config changes)."""
    global _suggestion_service
    with _suggestion_service_lock:
        _suggestion_service = None


ENV_SETTING_FIELDS: list[dict[str, Any]] = [
    # Zabbix
    {
        "key": "ZABBIX_API_URL",
        "label": "Zabbix API URL",
        "secret": False,
        "placeholder": "https://zabbix.example.com/api_jsonrpc.php",
        "category": "zabbix",
    },
    {
        "key": "ZABBIX_HOST",
        "label": "Zabbix Host",
        "secret": False,
        "placeholder": "https://zabbix.example.com",
        "category": "zabbix",
    },
    {
        "key": "ZABBIX_WEB_URL",
        "label": "Zabbix Web URL",
        "secret": False,
        "placeholder": "https://zabbix.example.com",
        "category": "zabbix",
    },
    {
        "key": "ZABBIX_API_TOKEN",
        "label": "Zabbix API Token",
        "secret": True,
        "placeholder": "paste-token-here",
        "category": "zabbix",
    },
    {
        "key": "ZABBIX_SEVERITIES",
        "label": "Zabbix Severities",
        "secret": False,
        "placeholder": "2,3,4",
        "category": "zabbix",
    },
    {
        "key": "ZABBIX_GROUP_ID",
        "label": "Zabbix Group ID",
        "secret": False,
        "placeholder": "optional",
        "category": "zabbix",
    },
    # NetBox & Atlassian
    {
        "key": "NETBOX_URL",
        "label": "NetBox URL",
        "secret": False,
        "placeholder": "https://netbox.example.com",
        "category": "net-atlassian",
    },
    {
        "key": "NETBOX_TOKEN",
        "label": "NetBox API Token",
        "secret": True,
        "placeholder": "paste-token-here",
        "category": "net-atlassian",
    },
    {
        "key": "NETBOX_DEBUG",
        "label": "NetBox Debug Logging",
        "secret": False,
        "placeholder": "0",
        "category": "net-atlassian",
    },
    {
        "key": "NETBOX_EXTRA_HEADERS",
        "label": "NetBox Extra Headers",
        "secret": False,
        "placeholder": "Key=Value;Other=Value",
        "category": "net-atlassian",
    },
    {
        "key": "NETBOX_DATA_DIR",
        "label": "NetBox Data Directory",
        "secret": False,
        "placeholder": "data",
        "category": "net-atlassian",
    },
    {
        "key": "ATLASSIAN_BASE_URL",
        "label": "Atlassian Base URL",
        "secret": False,
        "placeholder": "https://your-domain.atlassian.net",
        "category": "net-atlassian",
    },
    {
        "key": "ATLASSIAN_EMAIL",
        "label": "Atlassian Email",
        "secret": False,
        "placeholder": "user@example.com",
        "category": "net-atlassian",
    },
    {
        "key": "ATLASSIAN_API_TOKEN",
        "label": "Atlassian API Token",
        "secret": True,
        "placeholder": "paste-token-here",
        "category": "net-atlassian",
    },
    {
        "key": "CONFLUENCE_CMDB_PAGE_ID",
        "label": "Confluence CMDB Page ID",
        "secret": False,
        "placeholder": "981533033",
        "category": "net-atlassian",
    },
    {
        "key": "CONFLUENCE_DEVICES_PAGE_ID",
        "label": "Confluence Devices Page ID",
        "secret": False,
        "placeholder": "optional",
        "category": "net-atlassian",
    },
    {
        "key": "CONFLUENCE_VMS_PAGE_ID",
        "label": "Confluence VMs Page ID",
        "secret": False,
        "placeholder": "optional",
        "category": "net-atlassian",
    },
    {
        "key": "CONFLUENCE_ENABLE_TABLE_FILTER",
        "label": "Enable Table Filter Macro",
        "secret": False,
        "placeholder": "0 or 1",
        "category": "net-atlassian",
    },
    {
        "key": "CONFLUENCE_ENABLE_TABLE_SORT",
        "label": "Enable Table Sort Macro",
        "secret": False,
        "placeholder": "0 or 1",
        "category": "net-atlassian",
    },
    # Chat providers
    {"key": "OPENAI_API_KEY", "label": "OpenAI API Key", "secret": True, "placeholder": "sk-...", "category": "chat"},
    {
        "key": "CHAT_DEFAULT_MODEL_OPENAI",
        "label": "OpenAI Default Model",
        "secret": False,
        "placeholder": "gpt-4o-mini",
        "category": "chat",
    },
    {
        "key": "OPENROUTER_API_KEY",
        "label": "OpenRouter API Key",
        "secret": True,
        "placeholder": "or-...",
        "category": "chat",
    },
    {
        "key": "CHAT_DEFAULT_MODEL_OPENROUTER",
        "label": "OpenRouter Default Model",
        "secret": False,
        "placeholder": "openrouter/auto",
        "category": "chat",
    },
    {
        "key": "ANTHROPIC_API_KEY",
        "label": "Anthropic API Key",
        "secret": True,
        "placeholder": "api-key",
        "category": "chat",
    },
    {
        "key": "CHAT_DEFAULT_MODEL_CLAUDE",
        "label": "Anthropic Default Model",
        "secret": False,
        "placeholder": "claude-3-5-sonnet",
        "category": "chat",
    },
    {
        "key": "GOOGLE_API_KEY",
        "label": "Google (Gemini) API Key",
        "secret": True,
        "placeholder": "AIza...",
        "category": "chat",
    },
    {
        "key": "CHAT_DEFAULT_MODEL_GEMINI",
        "label": "Gemini Default Model",
        "secret": False,
        "placeholder": "gemini-1.5-flash",
        "category": "chat",
    },
    {
        "key": "CHAT_DEFAULT_PROVIDER",
        "label": "Default Chat Provider",
        "secret": False,
        "placeholder": "openai",
        "category": "chat",
    },
    {
        "key": "CHAT_DEFAULT_TEMPERATURE",
        "label": "Default Temperature",
        "secret": False,
        "placeholder": "0.2",
        "category": "chat",
    },
    {
        "key": "CHAT_SYSTEM_PROMPT",
        "label": "System Instructions",
        "secret": False,
        "placeholder": "Optional system prompt",
        "category": "chat",
    },
    # Export & reporting
    {
        "key": "NETBOX_XLSX_ORDER_FILE",
        "label": "Column Order Template",
        "secret": False,
        "placeholder": "path/to/column_order.xlsx",
        "category": "export",
    },
    # API & Web UI
    {"key": "LOG_LEVEL", "label": "Log Level", "secret": False, "placeholder": "INFO", "category": "api"},
    {
        "key": "ATLAS_API_TOKEN",
        "label": "Atlas API Token",
        "secret": True,
        "placeholder": "optional",
        "category": "api",
    },
    {
        "key": "ATLAS_UI_PASSWORD",
        "label": "UI Password",
        "secret": True,
        "placeholder": "optional",
        "category": "api",
    },
    {
        "key": "ATLAS_UI_SECRET",
        "label": "UI Session Secret",
        "secret": True,
        "placeholder": "auto-generated if empty",
        "category": "api",
    },
    {
        "key": "ATLAS_SSL_CERTFILE",
        "label": "SSL Certificate File",
        "secret": False,
        "placeholder": "certs/localhost.pem",
        "category": "api",
    },
    {
        "key": "ATLAS_SSL_KEYFILE",
        "label": "SSL Key File",
        "secret": False,
        "placeholder": "certs/localhost-key.pem",
        "category": "api",
    },
    {
        "key": "ATLAS_SSL_KEY_PASSWORD",
        "label": "SSL Key Password",
        "secret": True,
        "placeholder": "optional",
        "category": "api",
    },
    {
        "key": "UI_THEME_DEFAULT",
        "label": "Default Theme",
        "secret": False,
        "placeholder": "silver-dark",
        "category": "api",
    },
    # Backup
    {"key": "BACKUP_ENABLE", "label": "Backup Enabled", "secret": False, "placeholder": "1", "category": "backup"},
    {
        "key": "BACKUP_TYPE",
        "label": "Backup Type",
        "secret": False,
        "placeholder": "local, sftp, scp, webdav",
        "category": "backup",
    },
    {
        "key": "BACKUP_ENCRYPTION_PASSWORD",
        "label": "Encryption Password",
        "secret": True,
        "placeholder": "secure-password-for-backups",
        "category": "backup",
    },
    {
        "key": "BACKUP_LOCAL_PATH",
        "label": "Local Backup Path",
        "secret": False,
        "placeholder": "backups",
        "category": "backup",
    },
    {
        "key": "BACKUP_HOST",
        "label": "SFTP/SCP Host",
        "secret": False,
        "placeholder": "backup.example.com",
        "category": "backup",
    },
    {"key": "BACKUP_PORT", "label": "SFTP/SCP Port", "secret": False, "placeholder": "22", "category": "backup"},
    {
        "key": "BACKUP_USERNAME",
        "label": "SFTP/SCP Username",
        "secret": False,
        "placeholder": "backup_user",
        "category": "backup",
    },
    {
        "key": "BACKUP_PASSWORD",
        "label": "SFTP/SCP Password",
        "secret": True,
        "placeholder": "password",
        "category": "backup",
    },
    {
        "key": "BACKUP_PRIVATE_KEY_PATH",
        "label": "SFTP/SCP Private Key Path",
        "secret": False,
        "placeholder": "~/.ssh/id_rsa",
        "category": "backup",
    },
    {
        "key": "BACKUP_REMOTE_PATH",
        "label": "SFTP/SCP Remote Path",
        "secret": False,
        "placeholder": "/backups/atlas",
        "category": "backup",
    },
    {
        "key": "BACKUP_WEBDAV_URL",
        "label": "WebDAV URL",
        "secret": False,
        "placeholder": "https://file.example.com/remote.php/webdav",
        "category": "backup",
    },
    {
        "key": "BACKUP_WEBDAV_USERNAME",
        "label": "WebDAV Username",
        "secret": False,
        "placeholder": "backup",
        "category": "backup",
    },
    {
        "key": "BACKUP_WEBDAV_PASSWORD",
        "label": "WebDAV Password",
        "secret": True,
        "placeholder": "password",
        "category": "backup",
    },
    {
        "key": "BACKUP_WEBDAV_PATH",
        "label": "WebDAV Folder",
        "secret": False,
        "placeholder": "Atlas",
        "category": "backup",
    },
    {
        "key": "BACKUP_CREATE_TIMESTAMPED_DIRS",
        "label": "Create Timestamped Directories",
        "secret": False,
        "placeholder": "true",
        "category": "backup",
    },
    # Suggestions / Airtable
    {
        "key": "SUGGESTIONS_BACKEND",
        "label": "Suggestions Backend",
        "secret": False,
        "placeholder": "csv or airtable",
        "category": "suggestions",
    },
    {
        "key": "AIRTABLE_PAT",
        "label": "Airtable Personal Access Token",
        "secret": True,
        "placeholder": "patXXX...",
        "category": "suggestions",
    },
    {
        "key": "AIRTABLE_BASE_ID",
        "label": "Airtable Base ID",
        "secret": False,
        "placeholder": "appXXXXXXXXXXXXXX",
        "category": "suggestions",
    },
    {
        "key": "AIRTABLE_TABLE_NAME",
        "label": "Airtable Table Name",
        "secret": False,
        "placeholder": "Suggestions",
        "category": "suggestions",
    },
    {
        "key": "AIRTABLE_CACHE_TTL",
        "label": "Airtable Cache TTL (seconds)",
        "secret": False,
        "placeholder": "300",
        "category": "suggestions",
    },
]


def _write_log(msg: str) -> None:
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with LOG_PATH.open("a", encoding="utf-8") as fh:
            fh.write(f"[{ts}] {msg.rstrip()}\n")
    except Exception:
        # Logging failures should never crash the API
        pass


def _require_classification(value: str | None) -> str:
    if value is None or str(value).strip() == "":
        return "Could have"
    raw = str(value).strip().lower()
    for name in SUGGESTION_CLASSIFICATIONS:
        if raw == name.lower():
            return name
    raise HTTPException(status_code=400, detail=f"Invalid classification: {value}")


def _require_status(value: str | None) -> str:
    if value is None or str(value).strip() == "":
        return "new"
    raw = str(value).strip().lower()
    if raw not in SUGGESTION_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status: {value}")
    return raw


def _env_file_path() -> Path:
    env_path = load_env()
    if not env_path.exists():
        env_path.touch()
    return env_path


def _read_env_values() -> dict[str, str]:
    env_path = _env_file_path()
    values = dotenv_values(env_path)
    # dotenv_values returns OrderedDict[str, Optional[str]]
    clean: dict[str, str] = {}
    for key, value in values.items():
        if value is None:
            continue
        clean[str(key)] = str(value)
    return clean


def _read_env_defaults() -> dict[str, str]:
    example_path = project_root() / ".env.example"
    if not example_path.exists():
        return {}
    values = dotenv_values(example_path)
    return {str(k): str(v) for k, v in values.items() if v is not None}


def _write_env_value(key: str, value: str | None) -> None:
    key = key.strip()
    env_path = _env_file_path()
    lines = env_path.read_text().splitlines() if env_path.exists() else []
    written = False
    new_lines: list[str] = []
    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in line:
            new_lines.append(line)
            continue
        current_key, _, _ = line.partition("=")
        if current_key.strip() == key:
            if value is None:
                written = True
                continue  # remove line
            new_lines.append(f"{key}={value}")
            written = True
        else:
            new_lines.append(line)
    if not written and value is not None:
        if new_lines and new_lines[-1] != "":
            new_lines.append("")
        new_lines.append(f"{key}={value}")
    env_path.write_text("\n".join(new_lines).rstrip() + "\n")
    load_env(override=True)


class SuggestionCreate(BaseModel):
    title: str
    summary: str | None = ""
    classification: str | None = None


class SuggestionUpdate(BaseModel):
    title: str | None = None
    summary: str | None = None
    classification: str | None = None
    status: str | None = None


class SuggestionLikeRequest(BaseModel):
    delta: int = 1


class SuggestionCommentCreate(BaseModel):
    text: str


class EnvUpdateRequest(BaseModel):
    key: str
    value: str | None = None


class EnvResetRequest(BaseModel):
    key: str


# Include modular routes from interfaces
app.include_router(bootstrap_api())

# Include monitoring routes
try:
    from infrastructure_atlas.interfaces.api.routes.monitoring import router as monitoring_router

    app.include_router(monitoring_router)
except ImportError:
    logger.warning("Monitoring routes not available - optional dependencies missing")

# Re-export functions for CLI compatibility (these live in route modules but CLI imports from app.py)
from infrastructure_atlas.interfaces.api.routes.jira import jira_search  # noqa: E402, F401
from infrastructure_atlas.interfaces.api.routes.netbox import netbox_search  # noqa: E402, F401
from infrastructure_atlas.interfaces.api.routes.zabbix import zabbix_problems  # noqa: E402, F401

# Add authentication middleware
app.add_middleware(AuthMiddleware)

# Session support for UI auth (cookie-based). Added AFTER AuthMiddleware so
# that Session is applied first (outermost), making request.session available.
app.add_middleware(
    SessionMiddleware,
    secret_key=UI_SECRET,
    session_cookie=SESSION_COOKIE_NAME,
    same_site="lax",
)


@app.get("/suggestions")
def suggestions_list() -> dict:
    service = _get_suggestion_service()
    items = service.list_suggestions()
    meta = meta_to_dto(service.get_meta())
    dto = SuggestionListDTO(
        items=suggestions_to_dto(items),
        total=len(items),
        meta=meta,
    )
    return dto.dict_clean()


@app.get("/suggestions/config")
def suggestions_config() -> dict:
    """Return the current suggestions backend configuration."""
    backend_type = get_suggestion_backend_type()
    airtable_configured = bool(os.getenv("AIRTABLE_PAT", "").strip() and os.getenv("AIRTABLE_BASE_ID", "").strip())
    return {
        "backend": backend_type,
        "airtable_configured": airtable_configured,
        "airtable_base_id": os.getenv("AIRTABLE_BASE_ID", "") if airtable_configured else None,
        "airtable_table_name": os.getenv("AIRTABLE_TABLE_NAME", "Suggestions") if airtable_configured else None,
        "cache_ttl": int(os.getenv("AIRTABLE_CACHE_TTL", "300")),
    }


@app.post("/suggestions/migrate")
def suggestions_migrate() -> dict:
    """Migrate suggestions from CSV to Airtable."""
    csv_path = _data_dir() / "suggestions.csv"
    if not csv_path.exists():
        raise HTTPException(status_code=404, detail="No CSV file found to migrate")

    try:
        result = migrate_csv_to_airtable(csv_path)
        # Reset the service to pick up the new backend
        _reset_suggestion_service()
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Migration failed: {e}") from e


@app.get("/suggestions/{suggestion_id}")
def suggestions_detail(suggestion_id: str) -> dict:
    service = _get_suggestion_service()
    item = service.get_suggestion(suggestion_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Suggestion not found")
    meta = meta_to_dto(service.get_meta())
    dto = SuggestionItemDTO(
        item=suggestion_to_dto(item),
        meta=meta,
    )
    return dto.dict_clean()


@app.post("/suggestions")
def suggestions_create(req: SuggestionCreate) -> dict:
    title = (req.title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="Title is required")
    summary = (req.summary or "").strip()
    classification = _require_classification(req.classification)

    service = _get_suggestion_service()
    try:
        item = service.create_suggestion(title=title, summary=summary, classification=classification)
        dto = SuggestionItemDTO(item=suggestion_to_dto(item))
        return dto.dict_clean()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create suggestion: {e}") from e


@app.put("/suggestions/{suggestion_id}")
def suggestions_update(suggestion_id: str, req: SuggestionUpdate) -> dict:
    service = _get_suggestion_service()

    title = None
    if req.title is not None:
        new_title = req.title.strip()
        if not new_title:
            raise HTTPException(status_code=400, detail="Title cannot be empty")
        title = new_title

    summary = req.summary.strip() if req.summary else None if req.summary is None else ""
    classification = _require_classification(req.classification) if req.classification is not None else None
    status = _require_status(req.status) if req.status is not None else None

    try:
        item = service.update_suggestion(
            suggestion_id,
            title=title,
            summary=summary,
            classification=classification,
            status=status,
        )
        dto = SuggestionItemDTO(item=suggestion_to_dto(item))
        return dto.dict_clean()
    except SuggestionNotFoundError:
        raise HTTPException(status_code=404, detail="Suggestion not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update suggestion: {e}") from e


@app.post("/suggestions/{suggestion_id}/like")
def suggestions_like(suggestion_id: str, req: SuggestionLikeRequest) -> dict:
    delta = req.delta or 1
    service = _get_suggestion_service()

    try:
        item = service.like_suggestion(suggestion_id, delta=delta)
        dto = SuggestionItemDTO(item=suggestion_to_dto(item))
        return dto.dict_clean()
    except SuggestionNotFoundError:
        raise HTTPException(status_code=404, detail="Suggestion not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to like suggestion: {e}") from e


@app.post("/suggestions/{suggestion_id}/comments")
def suggestions_add_comment(suggestion_id: str, req: SuggestionCommentCreate) -> dict:
    text = (req.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Comment text is required")

    service = _get_suggestion_service()
    try:
        item, comment = service.add_comment(suggestion_id, text)
        dto = SuggestionCommentResponseDTO(
            item=suggestion_to_dto(item),
            comment=SuggestionCommentDTO.model_validate(comment),
        )
        return dto.dict_clean()
    except SuggestionNotFoundError:
        raise HTTPException(status_code=404, detail="Suggestion not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to add comment: {e}") from e


@app.delete("/suggestions/{suggestion_id}/comments/{comment_id}")
def suggestions_delete_comment(suggestion_id: str, comment_id: str) -> dict:
    service = _get_suggestion_service()

    try:
        item = service.delete_comment(suggestion_id, comment_id)
        dto = SuggestionItemDTO(item=suggestion_to_dto(item))
        return dto.dict_clean()
    except SuggestionNotFoundError as e:
        detail = "Comment not found" if "Comment" in str(e) else "Suggestion not found"
        raise HTTPException(status_code=404, detail=detail)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete comment: {e}") from e


@app.delete("/suggestions/{suggestion_id}")
def suggestions_delete(suggestion_id: str) -> dict:
    service = _get_suggestion_service()
    if not service.delete_suggestion(suggestion_id):
        raise HTTPException(status_code=404, detail="Suggestion not found")
    return {"ok": True}


def _build_admin_settings() -> list[dict[str, Any]]:
    env_values = _read_env_values()
    defaults = _read_env_defaults()
    settings: list[dict[str, Any]] = []
    for env_field in ENV_SETTING_FIELDS:
        key = env_field["key"]
        secret = bool(env_field.get("secret"))
        placeholder = env_field.get("placeholder") or ""
        category = env_field.get("category") or "general"
        current_value = os.getenv(key, "")
        has_value = bool(current_value)
        value = "" if secret else current_value
        placeholder_effective = placeholder
        if secret and has_value:
            placeholder_effective = "•••••• (hidden)"
        elif not placeholder_effective and not has_value:
            placeholder_effective = defaults.get(key, "")
        settings.append(
            {
                "key": key,
                "label": env_field.get("label", key),
                "secret": secret,
                "value": value,
                "has_value": has_value,
                "placeholder": placeholder,
                "placeholder_effective": placeholder_effective,
                "source": "file" if key in env_values else "env",
                "default": defaults.get(key, ""),
                "category": category,
            }
        )
    return settings


@app.get("/admin/env")
def admin_env_settings():
    settings = _build_admin_settings()
    backup_enabled = os.getenv("BACKUP_ENABLE", "1").strip().lower() not in {"0", "false", "no", "off"}
    backup_type = os.getenv("BACKUP_TYPE", "local").strip().lower()

    # Determine if backup is properly configured based on type
    backup_configured = False
    backup_target = ""

    if backup_type == "local":
        backup_path = os.getenv("BACKUP_LOCAL_PATH", "backups").strip()
        backup_configured = bool(backup_path)
        backup_target = backup_path
    elif backup_type in {"sftp", "scp"}:
        host = os.getenv("BACKUP_HOST", "").strip()
        username = os.getenv("BACKUP_USERNAME", "").strip()
        password = os.getenv("BACKUP_PASSWORD", "").strip()
        private_key = os.getenv("BACKUP_PRIVATE_KEY_PATH", "").strip()
        remote_path = os.getenv("BACKUP_REMOTE_PATH", "").strip()

        backup_configured = bool(host and username and (password or private_key))
        backup_target = (
            f"{username}@{host}:{remote_path}"
            if host and username and remote_path
            else f"{username}@{host}"
            if host and username
            else ""
        )

    settings_dto = [EnvSettingDTO.model_validate(setting) for setting in settings]
    backup_dto = BackupInfoDTO(
        enabled=backup_enabled,
        configured=backup_configured,
        type=backup_type,
        target=backup_target or None,
    )
    dto = AdminEnvResponseDTO(settings=settings_dto, backup=backup_dto)
    return dto.dict_clean()


@app.post("/admin/env")
def admin_env_update(req: EnvUpdateRequest):
    key = req.key.strip()
    if not key:
        raise HTTPException(status_code=400, detail="Key is required")
    field = next((f for f in ENV_SETTING_FIELDS if f["key"] == key), None)
    if field is None:
        raise HTTPException(status_code=404, detail="Unknown setting")
    if req.value is None:
        return admin_env_settings()
    _write_env_value(key, req.value)
    # Reset suggestion service if Airtable config changed
    if key.startswith("AIRTABLE_") or key == "SUGGESTIONS_BACKEND":
        _reset_suggestion_service()
    return admin_env_settings()


@app.post("/admin/env/reset")
def admin_env_reset(req: EnvResetRequest):
    key = req.key.strip()
    if not key:
        raise HTTPException(status_code=400, detail="Key is required")
    field = next((f for f in ENV_SETTING_FIELDS if f["key"] == key), None)
    if field is None:
        raise HTTPException(status_code=404, detail="Unknown setting")
    _write_env_value(key, None)
    # Reset suggestion service if Airtable config changed
    if key.startswith("AIRTABLE_") or key == "SUGGESTIONS_BACKEND":
        _reset_suggestion_service()
    return admin_env_settings()


@app.post("/admin/backup-sync")
def admin_backup_sync(user: OptionalUserDep = None):
    """Legacy backup sync endpoint - now uses new encrypted backup system."""
    actor = getattr(user, "username", None)
    with task_logging("admin.backup_sync", actor=actor, trigger="ui") as task_log:
        try:
            result = backup_manager.create_backup()
        except Exception as exc:  # pragma: no cover
            raise HTTPException(status_code=500, detail=str(exc)) from exc

        status = (result or {}).get("status")
        count = (result or {}).get("files_count")
        method = (result or {}).get("method")
        task_log.add_success(status=status, method=method, file_count=count)

        if status and status not in {"ok", "skipped"}:
            logger.warning(
                "Backup sync finished with non-OK status",
                extra={
                    "event": "task_warning",
                    "status": status,
                },
            )

        return result


class BackupCreateRequest(BaseModel):
    password: str | None = None


class BackupRestoreRequest(BaseModel):
    filename: str
    password: str | None = None
    dry_run: bool = False


@app.get("/admin/backup/config")
def admin_backup_config():
    """Get current backup configuration status."""
    backup_type = os.getenv("BACKUP_TYPE", "local").strip().lower()
    encryption_configured = bool(os.getenv("BACKUP_ENCRYPTION_PASSWORD", "").strip())

    config_status = {
        "type": backup_type,
        "encryption_configured": encryption_configured,
        "enabled": os.getenv("BACKUP_ENABLE", "1").strip().lower() not in {"0", "false", "no", "off"},
    }

    if backup_type == "local":
        config_status["local_path"] = os.getenv("BACKUP_LOCAL_PATH", "backups")
        config_status["configured"] = True
    elif backup_type in {"sftp", "scp"}:
        config_status["host"] = os.getenv("BACKUP_HOST", "")
        config_status["username"] = os.getenv("BACKUP_USERNAME", "")
        config_status["remote_path"] = os.getenv("BACKUP_REMOTE_PATH", "")
        config_status["configured"] = bool(config_status["host"] and config_status["username"])
    elif backup_type == "webdav":
        config_status["webdav_url"] = os.getenv("BACKUP_WEBDAV_URL", "")
        config_status["webdav_path"] = os.getenv("BACKUP_WEBDAV_PATH", "Atlas")
        config_status["configured"] = bool(config_status["webdav_url"])
    else:
        config_status["configured"] = False

    return config_status


@app.post("/admin/backup/create")
def admin_backup_create(req: BackupCreateRequest = None, user: OptionalUserDep = None):
    """Create an encrypted backup."""
    actor = getattr(user, "username", None)
    password = req.password if req else None

    with task_logging("admin.backup_create", actor=actor, trigger="ui") as task_log:
        try:
            result = backup_manager.create_backup(password=password)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc

        status = result.get("status")
        task_log.add_success(
            status=status,
            method=result.get("method"),
            file_count=result.get("files_count"),
            filename=result.get("filename"),
        )

        if status == "error":
            raise HTTPException(status_code=500, detail=result.get("reason", "Backup failed"))

        return result


@app.get("/admin/backup/list")
def admin_backup_list():
    """List available backups."""
    try:
        result = backup_manager.list_backups()
        if result.get("status") == "error":
            raise HTTPException(status_code=500, detail=result.get("reason", "Failed to list backups"))
        return result
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/admin/backup/restore")
def admin_backup_restore(req: BackupRestoreRequest, user: OptionalUserDep = None):
    """Restore a backup."""
    actor = getattr(user, "username", None)

    with task_logging("admin.backup_restore", actor=actor, trigger="ui") as task_log:
        try:
            result = backup_manager.restore_backup(
                filename=req.filename,
                password=req.password,
                dry_run=req.dry_run,
            )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc

        status = result.get("status")
        task_log.add_success(
            status=status,
            dry_run=req.dry_run,
            filename=req.filename,
            files_count=result.get("files_count"),
        )

        if status == "error":
            raise HTTPException(status_code=500, detail=result.get("reason", "Restore failed"))

        return result


@app.delete("/admin/backup/{filename}")
def admin_backup_delete(filename: str, user: OptionalUserDep = None):
    """Delete a backup."""
    actor = getattr(user, "username", None)

    with task_logging("admin.backup_delete", actor=actor, trigger="ui") as task_log:
        try:
            result = backup_manager.delete_backup(filename)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc

        task_log.add_success(filename=filename, status=result.get("status"))

        if result.get("status") == "error":
            raise HTTPException(status_code=404, detail=result.get("reason", "Backup not found"))

        return result


# ---------------------------
# Export endpoints (data only, no server-specific configs)
# ---------------------------


class ExportCreateRequest(BaseModel):
    password: str | None = None


class ExportImportRequest(BaseModel):
    filename: str
    password: str | None = None
    dry_run: bool = False


@app.post("/admin/export/create")
def admin_export_create(req: ExportCreateRequest = None, user: OptionalUserDep = None):
    """Create a data export (without server-specific configs like .env).

    Exports are safe to import on any server without overwriting its configuration.
    """
    actor = getattr(user, "username", None)
    password = req.password if req else None

    with task_logging("admin.export_create", actor=actor, trigger="ui") as task_log:
        try:
            result = backup_manager.create_export(password=password)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc

        status = result.get("status")
        task_log.add_success(
            status=status,
            filename=result.get("filename"),
            files_count=result.get("files_count"),
            type="export",
        )

        if status == "error":
            raise HTTPException(status_code=500, detail=result.get("reason", "Export failed"))

        return result


@app.get("/admin/export/list")
def admin_export_list():
    """List available exports."""
    try:
        result = backup_manager.list_exports()
        if result.get("status") == "error":
            raise HTTPException(status_code=500, detail=result.get("reason", "Failed to list exports"))
        return result
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/admin/export/import")
def admin_export_import(req: ExportImportRequest, user: OptionalUserDep = None):
    """Import a data export (without overwriting server-specific configs).

    This safely imports data from an export without affecting:
    - .env files (server-specific secrets)
    - Database files
    - Any server-specific configuration
    """
    actor = getattr(user, "username", None)

    with task_logging("admin.export_import", actor=actor, trigger="ui") as task_log:
        try:
            result = backup_manager.import_export(
                filename=req.filename,
                password=req.password,
                dry_run=req.dry_run,
            )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc

        status = result.get("status")
        task_log.add_success(
            status=status,
            dry_run=req.dry_run,
            filename=req.filename,
            files_count=result.get("files_count"),
            type="export",
        )

        if status == "error":
            raise HTTPException(status_code=500, detail=result.get("reason", "Import failed"))

        return result


@app.delete("/admin/export/{filename}")
def admin_export_delete(filename: str, user: OptionalUserDep = None):
    """Delete an export."""
    actor = getattr(user, "username", None)

    with task_logging("admin.export_delete", actor=actor, trigger="ui") as task_log:
        try:
            result = backup_manager.delete_export(filename)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc

        task_log.add_success(filename=filename, status=result.get("status"))

        if result.get("status") == "error":
            raise HTTPException(status_code=404, detail=result.get("reason", "Export not found"))

        return result


@app.get("/config/ui")
def ui_config():
    theme = os.getenv("UI_THEME_DEFAULT", "silver-dark").strip() or "silver-dark"
    return {"theme_default": theme}


def _build_sample_prompt(definition: ToolDefinition, args: Mapping[str, Any]) -> str:
    rendered_args = json.dumps(dict(args), indent=2, ensure_ascii=False, default=str) if args else "{}"
    return (
        f"Use the {definition.agent} agent's `{definition.key}` tool with these arguments:\n"
        f"{rendered_args}\n"
        "Provide a concise operational summary highlighting key metrics."
    )


@app.post("/tools/{tool_key}/sample")
def tools_run_sample(
    tool_key: str,
    user: OptionalUserDep,
    db: DbSessionDep,
    request: Request,
    payload: ToolSamplePayloadBody = None,
    provider: str | None = Query(None, description="Override the provider used for the sample run."),
    model: str | None = Query(None, description="Override the model used for the sample run."),
):
    # AI/LangChain functionality has been disabled
    raise HTTPException(
        status_code=501,
        detail="AI tool sampling functionality is currently disabled. LangChain dependencies have been removed.",
    )


# ---------------------------
# Zabbix integration (read-only)
# ---------------------------


# Serve favicon from project package location (png) as /favicon.ico
@app.get("/favicon.ico")
def favicon_ico():
    # Prefer png present at src/infrastructure_atlas/api/favicon.png
    png_path = Path(__file__).parent / "favicon.png"
    if png_path.exists():
        # Serve PNG under .ico path; browsers accept image/png
        return FileResponse(png_path, media_type="image/png")
    # Else, try static path under /app
    static_png = Path(__file__).parent / "static" / "favicon.png"
    if static_png.exists():
        return FileResponse(static_png, media_type="image/png")
    raise HTTPException(status_code=404, detail="favicon not found")


@app.get("/health")
def health():
    d = _data_dir()
    dev = _csv_path("netbox_devices_export.csv")
    vms = _csv_path("netbox_vms_export.csv")
    merged = _csv_path("netbox_merged_export.csv")
    return {
        "status": "ok",
        "data_dir": str(d),
        "files": {
            "devices_csv": dev.exists(),
            "vms_csv": vms.exists(),
            "merged_csv": merged.exists(),
        },
    }


@app.get("/column-order")
def column_order() -> list[str]:
    """Return preferred column order based on Systems CMDB.xlsx if available.

    Falls back to merged CSV headers if Excel not found; otherwise empty list.
    """
    try:
        # Prefer the Excel produced by merge step
        xlsx_path = _csv_path("Systems CMDB.xlsx")
        if xlsx_path.exists():
            try:
                from openpyxl import load_workbook

                wb = load_workbook(xlsx_path, read_only=True, data_only=True)
                ws = wb.worksheets[0]
                headers: list[str] = []
                for cell in ws[1]:
                    v = cell.value
                    if v is not None:
                        headers.append(str(v))
                wb.close()
                if headers:
                    return headers
            except Exception:
                pass
        # Fallback to merged CSV header
        csv_path = _csv_path("netbox_merged_export.csv")
        if csv_path.exists():
            with csv_path.open("r", encoding="utf-8") as fh:
                import csv as _csv

                reader = _csv.reader(fh)
                headers = next(reader, [])
                return [str(h) for h in headers if h]
    except Exception:
        pass
    return []


@app.get("/")
def root_redirect():
    # Serve the frontend at /app/
    return RedirectResponse(url="/app/")


# Mount static frontend
_static_dir = Path(__file__).parent / "static"


app.mount("/app", StaticFiles(directory=_static_dir, html=True), name="app")


@app.get("/logs/tail")
def logs_tail(n: int = Query(200, ge=1, le=5000)) -> dict:
    """
    Return the last N lines of the export log.
    Response: { "lines": ["..", ".."] }
    """
    if not LOG_PATH.exists():
        return {"lines": []}
    try:
        with LOG_PATH.open("r", encoding="utf-8", errors="ignore") as fh:
            lines = fh.readlines()[-n:]
        return {"lines": [ln.rstrip("\n") for ln in lines]}
    except Exception:
        return {"lines": []}

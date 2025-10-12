"""Application service orchestrating NetBox exports."""
from __future__ import annotations

import asyncio
import csv
import hashlib
import json
import os
from collections.abc import Callable, Iterable, Mapping
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from time import monotonic
from typing import Any, Protocol

from enreach_tools import backup_sync
from enreach_tools.application.exporter.netbox import (
    ExportPaths,
    LegacyScriptNetboxExporter,
    NativeNetboxExporter,
)
from enreach_tools.application.orchestration import JobHandler
from enreach_tools.domain.integrations import NetboxDeviceRecord, NetboxVMRecord
from enreach_tools.domain.tasks import JobPriority, JobSpec
from enreach_tools.infrastructure.external import NetboxClient, NetboxClientConfig
from enreach_tools.infrastructure.logging import get_logger, logging_context
from enreach_tools.infrastructure.metrics import record_netbox_export
from enreach_tools.infrastructure.tracing import span

try:  # optional dependencies for Excel export
    import pandas as pd  # type: ignore
    from openpyxl import Workbook, load_workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    from openpyxl.utils.dataframe import dataframe_to_rows  # type: ignore
    from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore
    EXCEL_AVAILABLE = True
except ImportError:  # pragma: no cover - optional dependency not installed
    EXCEL_AVAILABLE = False
    pd = None  # type: ignore
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore
    get_column_letter = None  # type: ignore
    dataframe_to_rows = None  # type: ignore
    Table = None  # type: ignore
    TableStyleInfo = None  # type: ignore


@dataclass(slots=True)
class CacheResourceSummary:
    resource: str
    total: int
    added: int
    updated: int
    removed: int


@dataclass(slots=True)
class NetboxCacheResult:
    generated_at: datetime
    path: Path
    summaries: Mapping[str, CacheResourceSummary]
    devices: list[NetboxDeviceRecord]
    vms: list[NetboxVMRecord]


class NetboxClientProtocol(Protocol):
    def list_devices(self, *, force_refresh: bool = False) -> Iterable[NetboxDeviceRecord]:
        ...

    def list_vms(self, *, force_refresh: bool = False) -> Iterable[NetboxVMRecord]:
        ...

    def list_device_metadata(self) -> Mapping[str, str]:
        ...

    def list_vm_metadata(self) -> Mapping[str, str]:
        ...

    def get_device(self, device_id: str | int) -> NetboxDeviceRecord:
        ...

    def get_devices_by_ids(self, identifiers: Iterable[int]) -> Iterable[NetboxDeviceRecord]:
        ...

    def get_vm(self, vm_id: str | int) -> NetboxVMRecord:
        ...

    def get_vms_by_ids(self, identifiers: Iterable[int]) -> Iterable[NetboxVMRecord]:
        ...


class NetboxExportService:
    JOB_NAME = "netbox.export.update"

    def __init__(self, *, client: NetboxClientProtocol, paths: ExportPaths) -> None:
        self._client = client
        self._paths = paths
        self._logger = get_logger(__name__)
        use_legacy = os.getenv("ENREACH_LEGACY_EXPORTER", "").strip().lower() in {"1", "true", "yes", "on"}
        if use_legacy:
            self._exporter = LegacyScriptNetboxExporter(paths)
        else:
            self._exporter = NativeNetboxExporter(client, paths, logger=self._logger)

    @classmethod
    def from_env(cls) -> NetboxExportService:
        url = os.environ.get("NETBOX_URL", "")
        token = os.environ.get("NETBOX_TOKEN", "")
        if not url or not token:
            raise ValueError("NETBOX_URL and NETBOX_TOKEN must be set")

        from enreach_tools.env import project_root

        data_dir_env = os.environ.get("NETBOX_DATA_DIR", "data")
        root = project_root()
        data_dir = Path(data_dir_env) if os.path.isabs(data_dir_env) else (root / data_dir_env)
        data_dir.mkdir(parents=True, exist_ok=True)

        paths = ExportPaths(
            data_dir=data_dir,
            devices_csv=data_dir / "netbox_devices_export.csv",
            vms_csv=data_dir / "netbox_vms_export.csv",
            merged_csv=data_dir / "netbox_merged_export.csv",
            excel_path=data_dir / "Systems CMDB.xlsx",
            scripts_root=root,
            manifest_path=data_dir / "netbox_export_manifest.json",
            cache_json=data_dir / "netbox_cache.json",
        )
        client = NetboxClient(NetboxClientConfig(url=url, token=token))
        return cls(client=client, paths=paths)

    def export_all(
        self,
        *,
        force: bool = False,
        verbose: bool = False,
        refresh_cache: bool = True,
    ) -> None:
        with logging_context(job=self.JOB_NAME, force=force, verbose=verbose), span(
            "netbox.export", job=self.JOB_NAME, force=force, verbose=verbose
        ):
            start = monotonic()
            status = "success"
            if verbose:
                self._logger.info("NetBox export started (verbose mode)")
            else:
                self._logger.info("NetBox export started")
            try:
                if refresh_cache:
                    cache_result = self.refresh_cache(force=force, verbose=verbose)
                else:
                    cache_result = self.load_cache()
                    if cache_result is None:
                        self._logger.info(
                            "NetBox cache not found; refreshing from API",
                            extra={"cache_file": self._paths.cache_json.as_posix()},
                        )
                        cache_result = self.refresh_cache(force=force, verbose=verbose)
                    else:
                        self._logger.debug(
                            "Using existing NetBox cache snapshot",
                            extra={"cache_file": cache_result.path.as_posix()},
                        )
                self._exporter.export(
                    force=force,
                    verbose=verbose,
                    devices=cache_result.devices,
                    vms=cache_result.vms,
                )
                self._merge_csv()
                self._create_excel()
                cache_invalidate = getattr(self._client, "invalidate_cache", None)
                if callable(cache_invalidate):
                    cache_invalidate()
            except Exception:
                status = "failure"
                duration = monotonic() - start
                record_netbox_export(duration_seconds=duration, mode="service", force=force, status=status)
                self._logger.exception(
                    "NetBox export failed",
                    extra={"duration_ms": int(duration * 1000)},
                )
                raise
            else:
                duration = monotonic() - start
                record_netbox_export(duration_seconds=duration, mode="service", force=force, status=status)
                self._logger.info(
                    "NetBox export completed",
                    extra={
                        "duration_ms": int(duration * 1000),
                        "devices_csv": self._paths.devices_csv.as_posix(),
                        "vms_csv": self._paths.vms_csv.as_posix(),
                        "merged_csv": self._paths.merged_csv.as_posix(),
                    },
                )

    async def export_all_async(
        self,
        *,
        force: bool = False,
        verbose: bool = False,
        refresh_cache: bool = True,
    ) -> None:
        await asyncio.to_thread(self.export_all, force=force, verbose=verbose, refresh_cache=refresh_cache)

    def refresh_cache(self, *, force: bool = False, verbose: bool = False) -> NetboxCacheResult:
        """Fetch NetBox resources and persist a JSON cache snapshot."""

        start = monotonic()
        status = "success"
        generated_at = datetime.now(UTC)
        summaries: dict[str, CacheResourceSummary] = {}
        previous_snapshot = self._load_cache_snapshot()
        with span("netbox.cache.refresh", job=self.JOB_NAME, force=force, verbose=verbose):
            try:
                prev_devices = self._cache_index(previous_snapshot, "devices")
                prev_vms = self._cache_index(previous_snapshot, "vms")
                metadata_devices = None if force else self._safe_fetch_metadata("device")
                metadata_vms = None if force else self._safe_fetch_metadata("vm")

                devices_map, device_records = self._reconcile_resource(
                    resource="devices",
                    metadata=metadata_devices,
                    previous=prev_devices,
                    force=force,
                    full_loader=lambda: self._client.list_devices(force_refresh=force),
                    partial_loader=getattr(self._client, "get_devices_by_ids", None),
                    single_loader=self._client.get_device,
                    cache_builder=self._device_cache_item,
                    cache_reader=self._device_from_cache_item,
                )
                vms_map, vm_records = self._reconcile_resource(
                    resource="vms",
                    metadata=metadata_vms,
                    previous=prev_vms,
                    force=force,
                    full_loader=lambda: self._client.list_vms(force_refresh=force),
                    partial_loader=getattr(self._client, "get_vms_by_ids", None),
                    single_loader=self._client.get_vm,
                    cache_builder=self._vm_cache_item,
                    cache_reader=self._vm_from_cache_item,
                )

                device_items = [self._ensure_item_hash(devices_map[key]) for key in self._sorted_keys(devices_map)]
                vm_items = [self._ensure_item_hash(vms_map[key]) for key in self._sorted_keys(vms_map)]
                snapshot = self._assemble_cache_snapshot(
                    generated_at=generated_at,
                    force=force,
                    device_items=device_items,
                    vm_items=vm_items,
                )
                self._write_cache_snapshot(snapshot)
                summaries = self._compute_cache_diff(previous_snapshot, snapshot)
            except Exception:
                status = "failure"
                duration = monotonic() - start
                record_netbox_export(duration_seconds=duration, mode="cache", force=force, status=status)
                self._logger.exception(
                    "Failed to refresh NetBox cache",
                    extra={"cache_file": self._paths.cache_json.as_posix()},
                )
                raise

        duration = monotonic() - start
        record_netbox_export(duration_seconds=duration, mode="cache", force=force, status=status)

        summary_payload = {
            name: {
                "total": summary.total,
                "added": summary.added,
                "updated": summary.updated,
                "removed": summary.removed,
            }
            for name, summary in summaries.items()
        }
        total_changes = sum(summary.added + summary.updated + summary.removed for summary in summaries.values())
        if total_changes:
            self._logger.info(
                "NetBox cache refreshed",
                extra={
                    "cache_file": self._paths.cache_json.as_posix(),
                    "resources": summary_payload,
                    "changes": total_changes,
                    "force": force,
                },
            )
        else:
            self._logger.debug(
                "NetBox cache refreshed with no changes",
                extra={"cache_file": self._paths.cache_json.as_posix(), "force": force},
            )

        return NetboxCacheResult(
            generated_at=generated_at,
            path=self._paths.cache_json,
            summaries=summaries,
            devices=device_records,
            vms=vm_records,
        )

    def build_job_spec(
        self,
        *,
        verbose: bool = False,
        force: bool = False,
        refresh_cache: bool = True,
        priority: JobPriority | None = None,
    ) -> JobSpec:
        """Create a `JobSpec` for scheduling a NetBox export run."""

        chosen_priority = priority or (JobPriority.HIGH if force else JobPriority.NORMAL)
        payload: Mapping[str, bool] = {"force": force, "verbose": verbose, "refresh_cache": refresh_cache}
        return JobSpec(name=self.JOB_NAME, payload=payload, priority=chosen_priority)

    def job_handler(self) -> JobHandler:
        """Return an async handler suitable for an `AsyncJobRunner`."""

        async def _handler(record) -> Mapping[str, str | bool]:
            force_flag = bool(record.payload.get("force", False))
            verbose_flag = bool(record.payload.get("verbose", False))
            refresh_flag = bool(record.payload.get("refresh_cache", True))
            await self.export_all_async(force=force_flag, verbose=verbose_flag, refresh_cache=refresh_flag)
            return {
                "force": force_flag,
                "verbose": verbose_flag,
                "refresh_cache": refresh_flag,
                "data_dir": self._paths.data_dir.as_posix(),
            }

        return _handler

    def _device_cache_item(self, record: NetboxDeviceRecord) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "id": int(record.id),
            "name": record.name,
            "status": record.status,
            "status_label": record.status_label,
            "role": record.role,
            "tenant": record.tenant,
            "tenant_group": record.tenant_group,
            "site": record.site,
            "location": record.location,
            "cluster": record.cluster,
            "primary_ip": record.primary_ip,
            "primary_ip4": record.primary_ip4,
            "primary_ip6": record.primary_ip6,
            "oob_ip": record.oob_ip,
            "tags": list(record.tags),
            "last_updated": self._isoformat(record.last_updated),
            "custom_fields": record.custom_fields,
            "manufacturer": record.manufacturer,
            "model": record.model,
            "rack": record.rack,
            "rack_unit": record.rack_unit,
            "serial": record.serial,
            "asset_tag": record.asset_tag,
            "site_group": record.site_group,
            "region": record.region,
            "description": record.description,
            "raw": record.raw,
        }
        payload["hash"] = self._hash_payload(payload)
        return payload

    def _vm_cache_item(self, record: NetboxVMRecord) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "id": int(record.id),
            "name": record.name,
            "status": record.status,
            "status_label": record.status_label,
            "role": record.role,
            "tenant": record.tenant,
            "tenant_group": record.tenant_group,
            "site": record.site,
            "location": record.location,
            "cluster": record.cluster,
            "primary_ip": record.primary_ip,
            "primary_ip4": record.primary_ip4,
            "primary_ip6": record.primary_ip6,
            "oob_ip": record.oob_ip,
            "tags": list(record.tags),
            "last_updated": self._isoformat(record.last_updated),
            "custom_fields": record.custom_fields,
            "platform": record.platform,
            "role_detail": record.role_detail,
            "description": record.description,
            "raw": record.raw,
        }
        payload["hash"] = self._hash_payload(payload)
        return payload

    def _load_cache_snapshot(self) -> dict[str, Any] | None:
        path = self._paths.cache_json
        if not path.exists():
            return None
        try:
            with path.open("r", encoding="utf-8") as fh:
                return json.load(fh)
        except Exception as exc:
            self._logger.warning(
                "Failed to load existing NetBox cache; starting fresh",
                extra={"cache_file": path.as_posix(), "error": str(exc)},
            )
            return None

    def _write_cache_snapshot(self, snapshot: Mapping[str, Any]) -> None:
        path = self._paths.cache_json
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_suffix(".tmp")
        with temp_path.open("w", encoding="utf-8") as fh:
            json.dump(snapshot, fh, indent=2, sort_keys=True, ensure_ascii=False)
            fh.write("\n")
        temp_path.replace(path)

    def _safe_fetch_metadata(self, resource: str) -> Mapping[str, str] | None:
        attr_name = f"list_{resource}_metadata"
        fetcher = getattr(self._client, attr_name, None)
        if not callable(fetcher):
            return None
        try:
            result = fetcher()
        except Exception as exc:
            self._logger.debug(
                "Failed to fetch NetBox %s metadata; falling back to full refresh",
                resource,
                extra={"error": str(exc)},
            )
            return None
        if not isinstance(result, Mapping):
            return None
        normalized: dict[str, str] = {}
        for key, value in result.items():
            normalized[str(key)] = value if value is not None else ""
        return normalized

    def _reconcile_resource(
        self,
        *,
        resource: str,
        metadata: Mapping[str, str] | None,
        previous: Mapping[str, Mapping[str, Any]],
        force: bool,
        full_loader: Callable[[], Iterable[Any]],
        partial_loader: Callable[[Iterable[int]], Iterable[Any]] | None,
        single_loader: Callable[[int], Any],
        cache_builder: Callable[[Any], dict[str, Any]],
        cache_reader: Callable[[Mapping[str, Any]], Any],
    ) -> tuple[dict[str, dict[str, Any]], list[Any]]:
        previous_map: dict[str, Mapping[str, Any]] = {str(key): value for key, value in previous.items()}
        if force or not metadata or not previous_map:
            records = list(full_loader())
            items_map = {str(record.id): cache_builder(record) for record in records}
            return items_map, records

        metadata_map: dict[str, str] = {str(key): value for key, value in metadata.items()}
        items_map: dict[str, dict[str, Any]] = {}
        record_map: dict[int, Any] = {}
        to_refresh: list[int] = []

        for key, current_meta in metadata_map.items():
            prev_item = previous_map.get(key)
            if prev_item is None:
                to_refresh.append(int(key))
                continue
            prev_last = self._normalize_iso(prev_item.get("last_updated"))
            curr_last = self._normalize_iso(current_meta)
            if prev_last == curr_last:
                cached_item = self._ensure_item_hash(prev_item)
                items_map[key] = cached_item
                record_map[int(key)] = cache_reader(prev_item)
            else:
                to_refresh.append(int(key))

        removed_keys = set(previous_map.keys()) - set(metadata_map.keys())
        fetch_targets = sorted(set(to_refresh))
        fetched_records: list[Any] = []
        if fetch_targets:
            loader = partial_loader if callable(partial_loader) else None
            if loader:
                try:
                    fetched_records = list(loader(fetch_targets))
                except Exception as exc:
                    self._logger.debug(
                        "Partial fetch for NetBox %s failed; falling back to individual requests",
                        resource,
                        extra={"error": str(exc)},
                    )
                    fetched_records = []
            fetched_ids = {int(getattr(record, "id", -1)) for record in fetched_records}
            for identifier in fetch_targets:
                if identifier in fetched_ids:
                    continue
                try:
                    fetched_records.append(single_loader(identifier))
                except Exception:
                    continue

        for record in fetched_records:
            key = str(record.id)
            items_map[key] = cache_builder(record)
            record_map[int(record.id)] = record

        for removed in removed_keys:
            items_map.pop(removed, None)
            try:
                record_map.pop(int(removed), None)
            except ValueError:
                continue

        for key in metadata_map.keys():
            if key in items_map:
                continue
            prev_item = previous_map.get(key)
            if prev_item is None:
                continue
            cached_item = self._ensure_item_hash(prev_item)
            items_map[key] = cached_item
            record_map[int(key)] = cache_reader(prev_item)

        records_sorted = [record_map[idx] for idx in sorted(record_map)]
        return items_map, records_sorted

    def _assemble_cache_snapshot(
        self,
        *,
        generated_at: datetime,
        force: bool,
        device_items: list[dict[str, Any]],
        vm_items: list[dict[str, Any]],
    ) -> dict[str, Any]:
        return {
            "version": 2,
            "generated_at": self._isoformat(generated_at),
            "force": bool(force),
            "resources": {
                "devices": {
                    "count": len(device_items),
                    "items": device_items,
                },
                "vms": {
                    "count": len(vm_items),
                    "items": vm_items,
                },
            },
        }

    def _ensure_item_hash(self, item: Mapping[str, Any]) -> dict[str, Any]:
        data = dict(item)
        existing = data.get("hash")
        if not existing:
            data["hash"] = self._hash_payload(data)
        return data

    @staticmethod
    def _sorted_keys(mapping: Mapping[str, Any]) -> list[str]:
        try:
            return sorted(mapping, key=lambda key: int(key))
        except Exception:
            return sorted(mapping)

    def _device_from_cache_item(self, item: Mapping[str, Any]) -> NetboxDeviceRecord:
        tags = item.get("tags") or ()
        return NetboxDeviceRecord(
            id=int(item.get("id") or 0),
            name=str(item.get("name") or ""),
            status=item.get("status"),
            status_label=item.get("status_label"),
            role=item.get("role"),
            tenant=item.get("tenant"),
            tenant_group=item.get("tenant_group"),
            site=item.get("site"),
            location=item.get("location"),
            tags=tuple(tags) if not isinstance(tags, tuple) else tags,
            last_updated=self._parse_iso(item.get("last_updated")),
            primary_ip=item.get("primary_ip"),
            primary_ip4=item.get("primary_ip4"),
            primary_ip6=item.get("primary_ip6"),
            oob_ip=item.get("oob_ip"),
            custom_fields=item.get("custom_fields") or {},
            raw=item.get("raw") or {},
            source=None,
            manufacturer=item.get("manufacturer"),
            model=item.get("model"),
            rack=item.get("rack"),
            rack_unit=item.get("rack_unit"),
            serial=item.get("serial"),
            asset_tag=item.get("asset_tag"),
            cluster=item.get("cluster"),
            site_group=item.get("site_group"),
            region=item.get("region"),
            description=item.get("description"),
        )

    def _vm_from_cache_item(self, item: Mapping[str, Any]) -> NetboxVMRecord:
        tags = item.get("tags") or ()
        return NetboxVMRecord(
            id=int(item.get("id") or 0),
            name=str(item.get("name") or ""),
            status=item.get("status"),
            status_label=item.get("status_label"),
            role=item.get("role"),
            tenant=item.get("tenant"),
            tenant_group=item.get("tenant_group"),
            site=item.get("site"),
            location=item.get("location"),
            tags=tuple(tags) if not isinstance(tags, tuple) else tags,
            last_updated=self._parse_iso(item.get("last_updated")),
            primary_ip=item.get("primary_ip"),
            primary_ip4=item.get("primary_ip4"),
            primary_ip6=item.get("primary_ip6"),
            oob_ip=item.get("oob_ip"),
            custom_fields=item.get("custom_fields") or {},
            raw=item.get("raw") or {},
            source=None,
            cluster=item.get("cluster"),
            role_detail=item.get("role_detail"),
            platform=item.get("platform"),
            description=item.get("description"),
        )

    @staticmethod
    def _parse_iso(value: Any) -> datetime | None:
        if not value:
            return None
        if isinstance(value, datetime):
            dt = value
        else:
            text = str(value).strip()
            if not text:
                return None
            if text.endswith("Z"):
                text = text[:-1] + "+00:00"
            try:
                dt = datetime.fromisoformat(text)
            except ValueError:
                return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        return dt.astimezone(UTC)

    def _normalize_iso(self, value: Any) -> str:
        parsed = self._parse_iso(value)
        if parsed is None:
            return ""
        normalized = parsed.replace(microsecond=0)
        return normalized.isoformat().replace("+00:00", "Z")

    def load_cache(self) -> NetboxCacheResult | None:
        snapshot = self._load_cache_snapshot()
        if not snapshot:
            return None
        generated_at = self._parse_iso(snapshot.get("generated_at")) or datetime.now(UTC)
        device_index = self._cache_index(snapshot, "devices")
        vm_index = self._cache_index(snapshot, "vms")
        device_items = [self._ensure_item_hash(device_index[key]) for key in self._sorted_keys(device_index)]
        vm_items = [self._ensure_item_hash(vm_index[key]) for key in self._sorted_keys(vm_index)]
        device_records = [self._device_from_cache_item(item) for item in device_items]
        vm_records = [self._vm_from_cache_item(item) for item in vm_items]
        summaries = {
            "devices": CacheResourceSummary(resource="devices", total=len(device_records), added=0, updated=0, removed=0),
            "vms": CacheResourceSummary(resource="vms", total=len(vm_records), added=0, updated=0, removed=0),
        }
        return NetboxCacheResult(
            generated_at=generated_at,
            path=self._paths.cache_json,
            summaries=summaries,
            devices=device_records,
            vms=vm_records,
        )

    def _compute_cache_diff(
        self,
        previous: Mapping[str, Any] | None,
        current: Mapping[str, Any],
    ) -> dict[str, CacheResourceSummary]:
        summaries: dict[str, CacheResourceSummary] = {}
        for resource in ("devices", "vms"):
            prev_index = self._cache_index(previous, resource)
            curr_index = self._cache_index(current, resource)
            prev_keys = set(prev_index.keys())
            curr_keys = set(curr_index.keys())
            added = curr_keys - prev_keys
            removed = prev_keys - curr_keys
            updated = {
                key
                for key in (curr_keys & prev_keys)
                if self._resolve_item_hash(curr_index.get(key)) != self._resolve_item_hash(prev_index.get(key))
            }
            summaries[resource] = CacheResourceSummary(
                resource=resource,
                total=len(curr_index),
                added=len(added),
                updated=len(updated),
                removed=len(removed),
            )
        return summaries

    @staticmethod
    def _cache_index(snapshot: Mapping[str, Any] | None, resource: str) -> dict[str, Mapping[str, Any]]:
        if not snapshot:
            return {}
        resources = snapshot.get("resources")
        if not isinstance(resources, Mapping):
            return {}
        payload = resources.get(resource)
        if not isinstance(payload, Mapping):
            return {}
        items = payload.get("items")
        if not isinstance(items, list):
            return {}
        index: dict[str, Mapping[str, Any]] = {}
        for item in items:
            if not isinstance(item, Mapping):
                continue
            identifier = item.get("id")
            if identifier is None:
                continue
            index[str(identifier)] = item
        return index

    @staticmethod
    def _hash_payload(payload: Mapping[str, Any]) -> str:
        serializable = dict(payload)
        serializable.pop("hash", None)
        encoded = json.dumps(
            serializable,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=False,
            default=str,
        )
        return hashlib.sha256(encoded.encode("utf-8")).hexdigest()

    def _resolve_item_hash(self, item: Mapping[str, Any] | None) -> str:
        if not item:
            return ""
        existing = item.get("hash")
        if isinstance(existing, str) and existing:
            return existing
        return self._hash_payload(item)

    @staticmethod
    def _isoformat(value: datetime | None) -> str | None:
        if value is None:
            return None
        if value.tzinfo is None:
            value = value.replace(tzinfo=UTC)
        normalized = value.astimezone(UTC).replace(microsecond=0)
        return normalized.isoformat().replace("+00:00", "Z")

    def _write_devices_csv(self, devices: Iterable[NetboxDeviceRecord]) -> None:
        with self._paths.devices_csv.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["id", "name", "last_updated"])
            for device in devices:
                last_updated = device.last_updated.isoformat() if device.last_updated else ""
                writer.writerow([device.id, device.name, last_updated])

    def _write_vms_csv(self, vms: Iterable[NetboxVMRecord]) -> None:
        with self._paths.vms_csv.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["id", "name", "last_updated"])
            for vm in vms:
                last_updated = vm.last_updated.isoformat() if vm.last_updated else ""
                writer.writerow([vm.id, vm.name, last_updated])

    def _merge_csv(self) -> None:
        devices_file = self._paths.devices_csv
        vms_file = self._paths.vms_csv
        output_file = self._paths.merged_csv

        if not devices_file.exists():
            raise FileNotFoundError(f"Devices CSV not found: {devices_file}")
        if not vms_file.exists():
            raise FileNotFoundError(f"VMs CSV not found: {vms_file}")

        self._logger.info(
            "Merging NetBox CSVs",
            extra={
                "devices_csv": devices_file.as_posix(),
                "vms_csv": vms_file.as_posix(),
                "output_csv": output_file.as_posix(),
            },
        )

        with devices_file.open(encoding="utf-8") as fh:
            devices_reader = csv.reader(fh)
            devices_headers = next(devices_reader)
        with vms_file.open(encoding="utf-8") as fh:
            vms_reader = csv.reader(fh)
            vms_headers = next(vms_reader)

        merged_headers = devices_headers.copy()
        for header in vms_headers:
            if header not in merged_headers:
                merged_headers.append(header)
        merged_headers.append("netbox_type")
        header_positions = {header: idx for idx, header in enumerate(merged_headers)}

        devices_count = 0
        vms_count = 0

        with output_file.open("w", newline="", encoding="utf-8") as outfile:
            writer = csv.writer(outfile)
            writer.writerow(merged_headers)

            with devices_file.open(encoding="utf-8") as fh:
                reader = csv.reader(fh)
                next(reader)
                for row in reader:
                    merged_row = ["" for _ in merged_headers]
                    for idx, header in enumerate(devices_headers):
                        value = row[idx] if idx < len(row) else ""
                        merged_row[header_positions[header]] = value
                    merged_row[header_positions["netbox_type"]] = "devices"
                    writer.writerow(merged_row)
                    devices_count += 1

            with vms_file.open(encoding="utf-8") as fh:
                reader = csv.reader(fh)
                next(reader)
                for row in reader:
                    merged_row = ["" for _ in merged_headers]
                    for idx, header in enumerate(vms_headers):
                        value = row[idx] if idx < len(row) else ""
                        merged_row[header_positions[header]] = value
                    merged_row[header_positions["netbox_type"]] = "vms"
                    writer.writerow(merged_row)
                    vms_count += 1

        total = devices_count + vms_count
        self._logger.info(
            "NetBox CSV merge completed",
            extra={
                "devices_processed": devices_count,
                "vms_processed": vms_count,
                "total_records": total,
            },
        )
        if output_file.exists():
            size = output_file.stat().st_size
            self._logger.debug(
                "Merged CSV size",
                extra={"bytes": size, "megabytes": round(size / 1024 / 1024, 2)},
            )

        try:
            backup_sync.sync_paths([output_file], note="netbox_merge_csv")
        except Exception:  # pragma: no cover - best-effort logging
            pass

    def _create_excel(self) -> None:
        if not EXCEL_AVAILABLE or pd is None or Workbook is None or dataframe_to_rows is None:
            self._logger.info("Skipping Excel export - required libraries not available")
            return
        csv_file = self._paths.merged_csv
        if not csv_file.exists():
            self._logger.info("Skipping Excel export - merged CSV not found", extra={"merged_csv": csv_file.as_posix()})
            return

        excel_file = self._paths.excel_path
        self._logger.info("Creating Excel export", extra={"excel_file": excel_file.as_posix()})
        df = pd.read_csv(csv_file)

        order_candidates = [
            os.getenv("NETBOX_XLSX_ORDER_FILE"),
            str(self._paths.scripts_root / "netbox-export" / "etc" / "column_order.xlsx"),
            str(self._paths.data_dir / "netbox_merged_export.xlsx"),
        ]
        order_file = next((p for p in order_candidates if p and os.path.exists(p)), None)
        if order_file:
            self._logger.info("Applying column order", extra={"order_file": order_file})
            desired_order = self._load_column_order_from_xlsx(Path(order_file))
            if desired_order:
                ordered_cols = [c for c in desired_order if c in df.columns]
                tail_cols = [c for c in df.columns if c not in ordered_cols]
                df = df[ordered_cols + tail_cols]
            else:
                self._logger.warning("Column order file did not yield headers; keeping CSV order", extra={"order_file": order_file})
        else:
            self._logger.debug("No column order template found; keeping CSV order")

        wb = Workbook()
        ws = wb.active
        ws.title = "NetBox Inventory"

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        num_cols = len(df.columns)
        end_col = get_column_letter(num_cols)
        if len(df) > 0:
            table_range = f"A1:{end_col}{len(df) + 1}"
            self._logger.debug("Creating Excel table", extra={"range": table_range})
            table = Table(displayName="NetBoxInventory", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        else:
            self._logger.info("No data rows; skipping Excel table creation")

        ws.freeze_panes = "B2"
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(excel_file)
        if excel_file.exists():
            size = excel_file.stat().st_size
            self._logger.info(
                "Excel export completed",
                extra={
                    "excel_file": excel_file.as_posix(),
                    "bytes": size,
                    "megabytes": round(size / 1024 / 1024, 2),
                },
            )

        try:
            backup_sync.sync_paths([excel_file], note="netbox_merge_excel")
        except Exception:  # pragma: no cover
            pass

    @staticmethod
    def _load_column_order_from_xlsx(order_file: Path) -> list[str]:
        if not EXCEL_AVAILABLE or load_workbook is None:
            return []
        try:
            wb = load_workbook(order_file, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            headers: list[str] = []
            for cell in ws[1]:
                if cell.value is None:
                    continue
                headers.append(str(cell.value))
            wb.close()
            return headers
        except Exception:
            return []


__all__ = ["ExportPaths", "NetboxClientProtocol", "NetboxExportService"]
